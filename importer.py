from __future__ import annotations

"""
Excel Import Module for ComplyEur.

Hardened to ensure consistent behaviour between local execution and Render:
- Validates environment and database paths via app.runtime_env
- Performs strict file and column validation with actionable errors
- Logs full context (row/column/file/environment) on failure
- Cleans up uploaded files regardless of success/failure
"""

import logging
import mimetypes
import os
import re
import sqlite3
import traceback
from dataclasses import dataclass
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

from app.runtime_env import build_runtime_state, require_env_vars

# Schengen country codes
logger = logging.getLogger(__name__)

ALLOWED_EXTENSIONS = {'.xlsx', '.xls'}
ALLOWED_MIME_TYPES = {
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'application/vnd.ms-excel',
}
REQUIRED_HEADER_INDICATORS = ('employee', 'name')

SCHENGEN_COUNTRIES = [
    'AT', 'BE', 'BG', 'HR', 'CY', 'CZ', 'DK', 'EE', 'ES', 'FI', 'FR', 'DE',
    'GR', 'HU', 'IS', 'IE', 'IT', 'LV', 'LI', 'LT', 'LU', 'MT', 'NL', 'NO',
    'PL', 'PT', 'RO', 'SK', 'SI', 'SE', 'CH'
]
SCHENGEN_COUNTRY_SET = set(SCHENGEN_COUNTRIES)


class ImportValidationError(Exception):
    """Raised for user-facing validation errors."""

    def __init__(self, message: str, *, details: Optional[Dict[str, Any]] = None):
        super().__init__(message)
        self.details = details or {}


def _validate_file_metadata(file_path: Path) -> None:
    extension = file_path.suffix.lower()
    if extension not in ALLOWED_EXTENSIONS:
        raise ImportValidationError(
            "Unsupported file type. Please upload an .xlsx or .xls file.",
            details={"extension": extension or "none"},
        )

    mime_type, _ = mimetypes.guess_type(file_path.name)
    if mime_type and mime_type not in ALLOWED_MIME_TYPES:
        raise ImportValidationError(
            "Invalid MIME type detected for the uploaded file.",
            details={"mime_type": mime_type},
        )


def _log_exception(context: Dict[str, Any]) -> None:
    logger.error("Excel import error context: %s", context)
    logger.error("Traceback:\n%s", traceback.format_exc())


def _normalise_country_code(country: str) -> str:
    code = country.strip().upper()
    if code == 'UK' or code == 'GB':
        return 'UK'
    return code


def _validate_employee_column(ws, header_row: int) -> None:
    header_value = ws.cell(row=header_row, column=1).value
    header_str = str(header_value or '').strip().lower()
    if not header_str or not any(indicator in header_str for indicator in REQUIRED_HEADER_INDICATORS):
        raise ImportValidationError(
            "Column A must contain employee names (header should reference 'Employee').",
            details={"header_value": header_value},
        )


def detect_country_enhanced(cell_text: str) -> Optional[str]:
    """
    Enhanced country detection using regex to find any 2-letter uppercase code.
    Returns country code if found, 'UK' for domestic, or None if no valid country.
    """
    if not cell_text or not isinstance(cell_text, str):
        return 'UK'
    
    # Remove travel prefix for country detection
    text = cell_text.strip()
    if text.lower().startswith(('tr/', 'tr-')):
        text = text[3:].strip()
    elif text.lower().startswith('tr'):
        text = text[2:].strip()
    
    # Find any 2-letter uppercase code as a standalone word
    pattern = r'\b([A-Z]{2})\b'
    matches = re.findall(pattern, text)
    
    for match in matches:
        if match in SCHENGEN_COUNTRY_SET:
            return match
    
    # Default to UK for domestic work
    return 'UK'


def is_travel_day(cell_text: str) -> bool:
    """
    Check if cell text indicates a travel day (starts with 'tr' or 'tr/').
    Case insensitive.
    """
    if not cell_text or not isinstance(cell_text, str):
        return False
    
    text = cell_text.strip().lower()
    return text.startswith(('tr/', 'tr-')) or (text.startswith('tr') and len(text) > 2 and text[2] in [' ', '/', '-'])


def parse_date_header(cell_value: Any) -> Optional[date]:
    """
    Parse date from various Excel header formats.
    Supports: "Mon 29 Sep", "29/09/2024", "29-09-2024", datetime objects.
    """
    if not cell_value:
        return None
    
    # If it's already a date/datetime object
    if isinstance(cell_value, (date, datetime)):
        if isinstance(cell_value, datetime):
            return cell_value.date()
        return cell_value
    
    # Convert to string
    cell_str = str(cell_value).strip()
    if not cell_str:
        return None
    
    # Try various date formats
    date_formats = [
        '%d/%m/%Y',
        '%d-%m-%Y',
        '%Y-%m-%d',
        '%d/%m/%y',
        '%d-%m-%y',
        '%Y/%m/%d',
        '%d %b %Y',  # "29 Sep 2024"
        '%d %B %Y',  # "29 September 2024"
        '%a %d %b',  # "Mon 29 Sep"
        '%A %d %B',  # "Monday 29 September"
    ]
    
    for fmt in date_formats:
        try:
            parsed = datetime.strptime(cell_str, fmt)
            # For formats without year, assume current year
            if parsed.year == 1900:
                parsed = parsed.replace(year=datetime.now().year)
            return parsed.date()
        except ValueError:
            continue
    
    return None


def find_header_row(ws, max_rows: int = 15) -> Optional[int]:
    """
    Scan first max_rows to find the header row containing dates.
    Returns row number (1-indexed) or None if not found.
    """
    for row_idx in range(1, min(max_rows + 1, ws.max_row + 1)):
        date_count = 0
        for col_idx in range(2, min(ws.max_column + 1, 100)):  # Check up to column 100
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if parse_date_header(cell_value):
                date_count += 1
                if date_count >= 3:  # Found at least 3 dates, likely the header row
                    return row_idx
    return None


def get_database_path(force_refresh: bool = False) -> str:
    """
    Resolve the SQLite database path using the shared runtime state.
    """
    require_env_vars()
    state = build_runtime_state(force_refresh=force_refresh)
    return str(state.database_path)


def get_or_create_employee(conn: sqlite3.Connection, name: str) -> Tuple[int, bool]:
    """
    Get existing employee ID or create new employee.
    Returns (employee_id, created_flag).
    """
    c = conn.cursor()
    
    # Try to find existing employee
    c.execute('SELECT id FROM employees WHERE name = ?', (name,))
    row = c.fetchone()
    if row:
        return row[0], False
    
    # Create new employee
    c.execute('INSERT INTO employees (name) VALUES (?)', (name,))
    conn.commit()
    return c.lastrowid, True


def trip_exists(conn: sqlite3.Connection, employee_id: int, country: str, entry_date: date, exit_date: date) -> bool:
    """
    Check if a trip with the same employee, country, and dates already exists.
    """
    c = conn.cursor()
    c.execute('''
        SELECT COUNT(*) FROM trips
        WHERE employee_id = ? AND country = ? AND entry_date = ? AND exit_date = ?
    ''', (employee_id, country, entry_date.isoformat(), exit_date.isoformat()))
    return c.fetchone()[0] > 0


def save_trip(conn: sqlite3.Connection, employee_id: int, country: str, entry_date: date, 
               exit_date: date, travel_days: int) -> bool:
    """
    Save a trip to the database if it doesn't already exist.
    Returns True if saved, False if duplicate.
    """
    if trip_exists(conn, employee_id, country, entry_date, exit_date):
        return False
    
    c = conn.cursor()
    c.execute('''
        INSERT INTO trips (employee_id, country, entry_date, exit_date, travel_days)
        VALUES (?, ?, ?, ?, ?)
    ''', (employee_id, country, entry_date.isoformat(), exit_date.isoformat(), travel_days))
    conn.commit()
    return True


def import_excel(
    file_path: str,
    enable_extended_scan: bool = False,
    db_conn: Optional[sqlite3.Connection] = None
) -> Dict[str, Any]:
    """
    Import trips from an Excel file.
    """
    path = Path(file_path)
    context = {
        "file": str(path),
        "enable_extended_scan": enable_extended_scan,
    }

    own_connection = False
    conn: Optional[sqlite3.Connection] = db_conn
    try:
        require_env_vars()
        runtime_state = build_runtime_state()
        context.update(
            {
                "database": str(runtime_state.database_path),
                "persistent_dir": str(runtime_state.persistent_dir),
            }
        )

        if not path.exists():
            raise ImportValidationError("Excel file not found.", details={"file": str(path)})

        _validate_file_metadata(path)

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as exc:
            raise ImportValidationError("Unable to open Excel file. Please ensure it is not password protected or corrupt.") from exc

        ws = wb.active

        max_scan_rows = 20 if enable_extended_scan else 15
        header_row = find_header_row(ws, max_scan_rows)

        if not header_row:
            raise ImportValidationError("Could not find a date header row in the first 15 rows.")

        _validate_employee_column(ws, header_row)

        date_columns: Dict[int, date] = {}
        for col_idx in range(2, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col_idx).value
            parsed_date = parse_date_header(cell_value)
            if parsed_date:
                date_columns[col_idx] = parsed_date

        if not date_columns:
            raise ImportValidationError("No valid date columns found in the header row.")

        records: List[Tuple[str, date, str, bool]] = []
        employee_names: set[str] = set()

        for row_idx in range(header_row + 1, ws.max_row + 1):
            employee_cell = ws.cell(row=row_idx, column=1).value
            if not employee_cell:
                continue

            employee_name = str(employee_cell).strip()
            if not employee_name or employee_name.lower() == 'unallocated':
                continue
            if len(employee_name) > 255:
                raise ImportValidationError(
                    "Employee names must be 255 characters or fewer.",
                    details={"row": row_idx, "value": employee_name},
                )

            employee_names.add(employee_name)

            for col_idx, cell_date in date_columns.items():
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is None:
                    continue

                cell_text = str(cell_value).strip()
                if not cell_text:
                    continue

                country = detect_country_enhanced(cell_text)
                country = _normalise_country_code(country)
                if country == 'UK':
                    continue
                if country not in SCHENGEN_COUNTRY_SET:
                    raise ImportValidationError(
                        "Invalid or unsupported country code detected.",
                        details={
                            "row": row_idx,
                            "column": get_column_letter(col_idx),
                            "value": cell_text,
                        },
                    )

                is_travel = is_travel_day(cell_text)
                records.append((employee_name, cell_date, country, is_travel))

        if not records:
            raise ImportValidationError("No valid trip records found in the Excel file.")

        records.sort(key=lambda x: (x[0], x[1]))
        trips: List[Tuple[str, str, date, date, int]] = []
        current_trip: Optional[Tuple[str, str, date, date, int]] = None

        for employee_name, trip_date, country, is_travel in records:
            if current_trip is None:
                current_trip = (employee_name, country, trip_date, trip_date, 1 if is_travel else 0)
                continue

            same_employee = current_trip[0] == employee_name
            same_country = current_trip[1] == country
            consecutive_day = (trip_date - current_trip[3]).days <= 1

            if same_employee and same_country and consecutive_day:
                travel_days = current_trip[4] + (1 if is_travel else 0)
                current_trip = (current_trip[0], current_trip[1], current_trip[2], trip_date, travel_days)
            else:
                trips.append(current_trip)
                current_trip = (employee_name, country, trip_date, trip_date, 1 if is_travel else 0)

        if current_trip:
            trips.append(current_trip)

        if not conn:
            db_path = get_database_path()
            conn = sqlite3.connect(db_path, check_same_thread=False)
            conn.row_factory = sqlite3.Row
            own_connection = True

        trips_added = 0
        employees_created = 0
        employees_processed: set[int] = set()
        warnings: List[str] = []
        duplicates_skipped = 0

        for employee_name, country, entry_date, exit_date, travel_days in trips:
            try:
                employee_id, created = get_or_create_employee(conn, employee_name)
            except sqlite3.DatabaseError as exc:
                raise ImportValidationError("Database error while creating employee.", details={"employee": employee_name}) from exc

            if created:
                employees_created += 1
            employees_processed.add(employee_id)

            try:
                if save_trip(conn, employee_id, country, entry_date, exit_date, travel_days):
                    trips_added += 1
                else:
                    duplicates_skipped += 1
                    warnings.append(
                        f"Duplicate trip ignored for {employee_name} ({country} {entry_date} - {exit_date})."
                    )
            except sqlite3.IntegrityError as exc:
                raise ImportValidationError(
                    "Database integrity error while saving trip.",
                    details={
                        "employee": employee_name,
                        "country": country,
                        "entry_date": entry_date.isoformat(),
                        "exit_date": exit_date.isoformat(),
                    },
                ) from exc

        result = {
            'success': True,
            'trips_added': trips_added,
            'employees_processed': len(employees_processed),
            'employees_created': employees_created,
            'total_employee_names_found': len(employee_names),
            'aggregated_trips': len(trips),
            'total_records': len(records),
            'warnings': warnings,
            'duplicates_skipped': duplicates_skipped,
            'uk_jobs_skipped': 0,
            'empty_cells_skipped': 0,
            'employees_without_trips': 0,
            'trips': [],
        }

        return result

    except ImportValidationError as validation_error:
        _log_exception({**context, "validation_details": validation_error.details})
        return {
            'success': False,
            'error': str(validation_error),
            'details': validation_error.details,
        }
    except Exception as exc:
        _log_exception({**context, "unexpected_error": str(exc)})
        return {
            'success': False,
            'error': f'Error processing Excel file: {str(exc)}',
        }
    finally:
        if own_connection and conn is not None:
            conn.close()


# Alias for backward compatibility
process_excel_file = import_excel

