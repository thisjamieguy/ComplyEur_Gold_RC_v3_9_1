from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, make_response, send_file
import hashlib
import sqlite3
from datetime import datetime, timedelta
from functools import wraps
import os
import re
import openpyxl
from werkzeug.utils import secure_filename
from config import load_config, get_session_lifetime, save_config
from app.services.hashing import Hasher
from app.services.audit import write_audit
from app.services.retention import purge_expired_trips, get_expired_trips, anonymize_employee
from app.services.dsar import create_dsar_export, delete_employee_data, rectify_employee_name, get_employee_data
from app.services.exports import export_trips_csv, export_employee_report_pdf, export_all_employees_report_pdf
import io
import csv
import zipfile

# Application Version
APP_VERSION = "1.0 (pre-security)"

app = Flask(__name__)
app.secret_key = os.urandom(24)
CONFIG = load_config()

# Configure static files
app.static_folder = 'static'
DATABASE = os.path.join(os.path.dirname(__file__), 'eu_tracker.db')

# Make version available to all templates
@app.context_processor
def inject_version():
    return dict(app_version=APP_VERSION)

# Session cookie hardening (essential-only cookie)
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Strict',
    SESSION_COOKIE_SECURE=bool(CONFIG.get('SESSION_COOKIE_SECURE', False))
)
app.permanent_session_lifetime = get_session_lifetime(CONFIG)

# Upload configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Ensure upload folder exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Schengen country code mapping
COUNTRY_CODE_MAPPING = {
    'AT': 'Austria',
    'BE': 'Belgium',
    'BG': 'Bulgaria',
    'HR': 'Croatia',
    'CY': 'Cyprus',
    'CZ': 'Czech Republic',
    'DK': 'Denmark',
    'EE': 'Estonia',
    'FI': 'Finland',
    'FR': 'France',
    'DE': 'Germany',
    'GR': 'Greece',
    'HU': 'Hungary',
    'IS': 'Iceland',
    'IE': 'Ireland',
    'IT': 'Italy',
    'LV': 'Latvia',
    'LI': 'Liechtenstein',
    'LT': 'Lithuania',
    'LU': 'Luxembourg',
    'MT': 'Malta',
    'NL': 'Netherlands',
    'NO': 'Norway',
    'PL': 'Poland',
    'PT': 'Portugal',
    'RO': 'Romania',
    'SK': 'Slovakia',
    'SI': 'Slovenia',
    'ES': 'Spain',
    'SE': 'Sweden',
    'CH': 'Switzerland'
}

# Reverse lookup
COUNTRY_NAME_TO_CODE = {v: k for k, v in COUNTRY_CODE_MAPPING.items()}

hasher = Hasher(CONFIG.get('PASSWORD_HASH_SCHEME', 'argon2'))
LOGIN_ATTEMPTS = {}
MAX_ATTEMPTS = 5
ATTEMPT_WINDOW_SECONDS = 600

@app.template_filter('uk_date')
def uk_date(value):
    """Format YYYY-MM-DD or datetime to DD-MM-YYYY for UK display."""
    try:
        if not value:
            return value
        if isinstance(value, datetime):
            return value.strftime('%d-%m-%Y')
        if isinstance(value, str):
            try:
                dt = datetime.strptime(value, '%Y-%m-%d')
                return dt.strftime('%d-%m-%Y')
            except ValueError:
                return value
        return value
    except Exception:
        return value

@app.template_filter('country_name')
def country_name(value):
    try:
        if not value:
            return value
        code = str(value).upper()
        return COUNTRY_CODE_MAPPING.get(code, code)
    except Exception:
        return value

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def check_password(stored_hash, password):
    return stored_hash == hash_password(password)

def verify_and_upgrade_password(stored_hash: str, candidate: str):
    """Verify password against stored hash, supporting legacy SHA256.
    Returns (is_valid: bool, upgraded_hash_or_None)."""
    try:
        if hasher.verify(stored_hash, candidate):
            return True, None
    except Exception:
        pass
    # legacy sha256
    try:
        if stored_hash == hashlib.sha256(candidate.encode()).hexdigest():
            try:
                return True, hasher.hash(candidate)
            except Exception:
                # If upgrade hashing is unavailable, still allow login without upgrading
                return True, None
    except Exception:
        pass
    return False, None

def init_db():
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    
    # Enable foreign keys
    c.execute('PRAGMA foreign_keys = ON')
    
    # Create tables
    c.execute('CREATE TABLE IF NOT EXISTS admin (id INTEGER PRIMARY KEY, password_hash TEXT NOT NULL)')
    c.execute('CREATE TABLE IF NOT EXISTS employees (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL UNIQUE)')
    c.execute('''CREATE TABLE IF NOT EXISTS trips (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        employee_id INTEGER NOT NULL,
        country TEXT NOT NULL,
        entry_date TEXT NOT NULL,
        exit_date TEXT NOT NULL,
        travel_days INTEGER DEFAULT 0,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE
    )''')
    
    # Lightweight migrations for legacy schemas
    try:
        c.execute("PRAGMA table_info(trips)")
        cols = {row[1] for row in c.fetchall()}
        if 'country' not in cols:
            c.execute('ALTER TABLE trips ADD COLUMN country TEXT')
            if 'country_code' in cols:
                c.execute('UPDATE trips SET country = country_code WHERE country IS NULL')
        c.execute("PRAGMA table_info(trips)")
        cols = {row[1] for row in c.fetchall()}
        if 'travel_days' not in cols:
            c.execute('ALTER TABLE trips ADD COLUMN travel_days INTEGER DEFAULT 0')
        c.execute("PRAGMA table_info(trips)")
        cols = {row[1] for row in c.fetchall()}
        if 'created_at' not in cols:
            c.execute("ALTER TABLE trips ADD COLUMN created_at TEXT DEFAULT (datetime('now'))")
    except Exception:
        pass
    
    # Create indexes for performance and GDPR compliance
    try:
        # Index on employee_id for faster joins and lookups
        c.execute('CREATE INDEX IF NOT EXISTS idx_trips_employee_id ON trips(employee_id)')
        # Index on exit_date for retention policy queries
        c.execute('CREATE INDEX IF NOT EXISTS idx_trips_exit_date ON trips(exit_date)')
        # Composite index on employee_id + entry_date for sorting/filtering
        c.execute('CREATE INDEX IF NOT EXISTS idx_trips_employee_entry ON trips(employee_id, entry_date)')
        # Index on entry_date for date range queries
        c.execute('CREATE INDEX IF NOT EXISTS idx_trips_entry_date ON trips(entry_date)')
    except Exception as e:
        print(f"Warning: Could not create indexes: {e}")
    
    # Initialize admin account if not exists
    c.execute('SELECT COUNT(*) FROM admin')
    if c.fetchone()[0] == 0:
        c.execute('INSERT INTO admin (password_hash) VALUES (?)', (hasher.hash('admin123'),))
    
    conn.commit()
    conn.close()

def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.before_request
def enforce_session_idle_timeout():
    try:
        if session.get('logged_in'):
            now = datetime.utcnow()
            last = session.get('last_activity')
            if last:
                last_dt = datetime.strptime(last, '%Y-%m-%dT%H:%M:%SZ')
                if now - last_dt > get_session_lifetime(CONFIG):
                    session.clear()
                    flash('Session expired. Please log in again.')
                    return redirect(url_for('login'))
            session['last_activity'] = now.strftime('%Y-%m-%dT%H:%M:%SZ')
    except Exception:
        pass

def calculate_eu_days(employee_id):
    conn = get_db()
    c = conn.cursor()
    today = datetime.now().date()
    start_date = today - timedelta(days=180)
    c.execute('SELECT entry_date, exit_date FROM trips WHERE employee_id = ? AND exit_date >= ?', (employee_id, start_date.strftime('%Y-%m-%d')))
    trips = c.fetchall()
    total_days = 0
    for trip in trips:
        entry = datetime.strptime(trip['entry_date'], '%Y-%m-%d').date()
        exit = datetime.strptime(trip['exit_date'], '%Y-%m-%d').date()
        effective_entry = max(entry, start_date)
        effective_exit = min(exit, today)
        if effective_entry <= effective_exit:
            total_days += (effective_exit - effective_entry).days + 1
    conn.close()
    return total_days

def calculate_eu_days_at_date(trips_for_employee, reference_date):
    """Calculate EU days used within the 180-day window ending at reference_date.
    trips_for_employee: list of dicts with 'entry_date' and 'exit_date' (YYYY-MM-DD)
    reference_date: datetime.date
    """
    window_start = reference_date - timedelta(days=180)
    total_days = 0
    for trip in trips_for_employee:
        entry = datetime.strptime(trip['entry_date'], '%Y-%m-%d').date()
        exit = datetime.strptime(trip['exit_date'], '%Y-%m-%d').date()
        effective_entry = max(entry, window_start)
        effective_exit = min(exit, reference_date)
        if effective_entry <= effective_exit:
            total_days += (effective_exit - effective_entry).days + 1
    return total_days

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def parse_schedule_excel_file(filepath):
    """Parse work schedule Excel file and extract EU trips based on country codes"""
    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active

        trips = []
        errors = []

        # Get employee names from first column (skip rows 4-9, and 68+)
        employees = {}
        for row_idx in range(10, min(68, sheet.max_row + 1)):  # Rows 10-67
            name_cell = sheet.cell(row=row_idx, column=1).value
            if name_cell and str(name_cell).strip() not in ['None', 'Unallocated', '']:
                employees[row_idx] = str(name_cell).strip()

        print(f"Found {len(employees)} employees in rows 10-67")

        # Get date headers from row 3 (starting from column 2)
        date_columns = {}
        for col_idx in range(2, sheet.max_column + 1):
            date_cell = sheet.cell(row=3, column=col_idx).value
            if date_cell and isinstance(date_cell, datetime):
                date_columns[col_idx] = date_cell.date()

        print(f"Found {len(date_columns)} date columns")

        # All possible EU/Schengen country codes
        eu_country_codes = list(COUNTRY_CODE_MAPPING.keys())

        # Process each employee's schedule
        for emp_row, emp_name in employees.items():
            current_country = None
            trip_start = None

            for col_idx in sorted(date_columns.keys()):
                date = date_columns[col_idx]
                cell_value = sheet.cell(row=emp_row, column=col_idx).value

                # Clean cell value
                assignment = None
                if cell_value:
                    assignment = str(cell_value).strip()
                    if assignment in ['None', '', ' ']:
                        assignment = None

                # Look for country codes in the assignment
                found_country = None
                if assignment:
                    assignment_upper = assignment.upper()
                    for country_code in eu_country_codes:
                        if country_code in assignment_upper:
                            found_country = country_code
                            break

                # If country changes, process the previous trip
                if found_country != current_country:
                    if current_country and trip_start:
                        # End current trip (yesterday)
                        trip_end = date_columns.get(col_idx - 1, date)

                        trips.append({
                            'name': emp_name,
                            'country': COUNTRY_CODE_MAPPING[current_country],
                            'entry_date': trip_start.strftime('%Y-%m-%d'),
                            'exit_date': trip_end.strftime('%Y-%m-%d'),
                            'row': emp_row
                        })

                    # Start new trip if we found a country
                    current_country = found_country
                    trip_start = date if found_country else None

            # Handle trip that extends to end of schedule
            if current_country and trip_start:
                last_date = max(date_columns.values())
                trips.append({
                    'name': emp_name,
                    'country': COUNTRY_CODE_MAPPING[current_country],
                    'entry_date': trip_start.strftime('%Y-%m-%d'),
                    'exit_date': last_date.strftime('%Y-%m-%d'),
                    'row': emp_row
                })

        return trips, errors

    except Exception as e:
        return [], [f"Error reading schedule Excel file: {str(e)}"]

def parse_excel_file(filepath):
    """Parse Excel file and extract trip data - detects format automatically"""
    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active

        # Detect if this is a schedule format vs. regular trip data format
        # Schedule format has: dates in row 3, employee names starting around row 10
        is_schedule_format = False

        # Check if row 3 has dates
        row3_has_dates = False
        for col_idx in range(2, min(10, sheet.max_column + 1)):
            cell_value = sheet.cell(row=3, column=col_idx).value
            if isinstance(cell_value, datetime):
                row3_has_dates = True
                break

        # Check if we have employee names around row 10-20
        has_employee_names = False
        for row_idx in range(10, min(21, sheet.max_row + 1)):
            name_cell = sheet.cell(row=row_idx, column=1).value
            if name_cell and isinstance(name_cell, str) and len(name_cell.strip()) > 3:
                has_employee_names = True
                break

        if row3_has_dates and has_employee_names:
            print("Detected schedule format - using schedule parser")
            return parse_schedule_excel_file(filepath)
        else:
            print("Detected standard format - using standard parser")
            # Continue with standard parsing logic

        trips = []
        errors = []

        # Try to detect headers
        headers = {}
        for col_idx, cell in enumerate(sheet[1], 1):
            if cell.value:
                header = str(cell.value).lower().strip()
                if 'name' in header or 'employee' in header:
                    headers['name'] = col_idx
                elif 'country' in header or 'destination' in header:
                    headers['country'] = col_idx
                elif 'entry' in header or 'start' in header or 'arrival' in header:
                    headers['entry_date'] = col_idx
                elif 'exit' in header or 'end' in header or 'departure' in header:
                    headers['exit_date'] = col_idx

        # If headers not found, assume standard order: Name, Country, Entry, Exit
        if not headers:
            headers = {'name': 1, 'country': 2, 'entry_date': 3, 'exit_date': 4}

        # Process data rows (starting from row 2)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):  # Skip empty rows
                continue

            try:
                # Extract data based on detected headers
                name = str(row[headers.get('name', 1) - 1] or '').strip() if len(row) >= headers.get('name', 1) else ''
                country_code = str(row[headers.get('country', 2) - 1] or '').strip().upper() if len(row) >= headers.get('country', 2) else ''
                entry_date = row[headers.get('entry_date', 3) - 1] if len(row) >= headers.get('entry_date', 3) else None
                exit_date = row[headers.get('exit_date', 4) - 1] if len(row) >= headers.get('exit_date', 4) else None

                # Validate required fields
                if not name:
                    errors.append(f"Row {row_idx}: Missing employee name")
                    continue

                if not country_code:
                    errors.append(f"Row {row_idx}: Missing country code")
                    continue

                # Convert country code to full name
                country_name = COUNTRY_CODE_MAPPING.get(country_code)
                if not country_name:
                    errors.append(f"Row {row_idx}: Unknown country code '{country_code}'")
                    continue

                # Parse dates
                if isinstance(entry_date, datetime):
                    entry_date_str = entry_date.strftime('%Y-%m-%d')
                elif isinstance(entry_date, str):
                    try:
                        parsed_date = datetime.strptime(entry_date.strip(), '%Y-%m-%d')
                        entry_date_str = parsed_date.strftime('%Y-%m-%d')
                    except ValueError:
                        try:
                            parsed_date = datetime.strptime(entry_date.strip(), '%d/%m/%Y')
                            entry_date_str = parsed_date.strftime('%Y-%m-%d')
                        except ValueError:
                            errors.append(f"Row {row_idx}: Invalid entry date format '{entry_date}'")
                            continue
                else:
                    errors.append(f"Row {row_idx}: Missing entry date")
                    continue

                if isinstance(exit_date, datetime):
                    exit_date_str = exit_date.strftime('%Y-%m-%d')
                elif isinstance(exit_date, str):
                    try:
                        parsed_date = datetime.strptime(exit_date.strip(), '%Y-%m-%d')
                        exit_date_str = parsed_date.strftime('%Y-%m-%d')
                    except ValueError:
                        try:
                            parsed_date = datetime.strptime(exit_date.strip(), '%d/%m/%Y')
                            exit_date_str = parsed_date.strftime('%Y-%m-%d')
                        except ValueError:
                            errors.append(f"Row {row_idx}: Invalid exit date format '{exit_date}'")
                            continue
                else:
                    errors.append(f"Row {row_idx}: Missing exit date")
                    continue

                # Validate date logic
                if entry_date_str >= exit_date_str:
                    errors.append(f"Row {row_idx}: Entry date must be before exit date")
                    continue

                trips.append({
                    'name': name,
                    'country': country_name,
                    'entry_date': entry_date_str,
                    'exit_date': exit_date_str,
                    'row': row_idx
                })

            except Exception as e:
                errors.append(f"Row {row_idx}: Error processing row - {str(e)}")

        return trips, errors

    except Exception as e:
        return [], [f"Error reading Excel file: {str(e)}"]

def find_employee_by_name(name):
    """Find employee by name (case-insensitive)"""
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT id, name FROM employees WHERE LOWER(name) = LOWER(?)', (name,))
    result = c.fetchone()
    conn.close()
    return result

def trip_exists(employee_id, entry_date, exit_date):
    """Check if trip already exists"""
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT id FROM trips WHERE employee_id = ? AND entry_date = ? AND exit_date = ?',
              (employee_id, entry_date, exit_date))
    result = c.fetchone()
    conn.close()
    return result is not None

@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/privacy')
def privacy():
    return render_template('privacy.html',
                         current_date=datetime.now().strftime('%Y-%m-%d'),
                         retention_months=CONFIG.get('RETENTION_MONTHS', 36),
                         session_timeout_minutes=CONFIG.get('SESSION_IDLE_TIMEOUT_MINUTES', 30),
                         max_login_attempts=MAX_ATTEMPTS,
                         rate_limit_window=ATTEMPT_WINDOW_SECONDS // 60)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        ip = request.remote_addr or 'local'
        now = datetime.utcnow()
        attempts = LOGIN_ATTEMPTS.get(ip)
        if attempts:
            count, first_ts = attempts
            if (now - first_ts).total_seconds() <= ATTEMPT_WINDOW_SECONDS and count >= MAX_ATTEMPTS:
                write_audit(CONFIG['AUDIT_LOG_PATH'], 'login_rate_limited', 'admin', {'ip': ip})
                return render_template('login.html', error='Too many attempts. Try again later.')
            if (now - first_ts).total_seconds() > ATTEMPT_WINDOW_SECONDS:
                LOGIN_ATTEMPTS[ip] = (0, now)
        else:
            LOGIN_ATTEMPTS[ip] = (0, now)

        password = request.form.get('password', '')
        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT password_hash FROM admin WHERE id = 1')
        admin = c.fetchone()
        if not admin:
            conn.close()
            return render_template('login.html', error='Setup incomplete')
        ok, new_hash = verify_and_upgrade_password(admin['password_hash'], password)
        if ok:
            if new_hash:
                c.execute('UPDATE admin SET password_hash = ? WHERE id = 1', (new_hash,))
                conn.commit()
            conn.close()
            session['logged_in'] = True
            session['last_activity'] = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')
            write_audit(CONFIG['AUDIT_LOG_PATH'], 'login_success', 'admin', {'ip': ip})
            return redirect(url_for('dashboard'))
        # failed
        conn.close()
        count, first_ts = LOGIN_ATTEMPTS.get(ip, (0, now))
        LOGIN_ATTEMPTS[ip] = (count + 1, first_ts)
        write_audit(CONFIG['AUDIT_LOG_PATH'], 'login_failure', 'admin', {'ip': ip})
        return render_template('login.html', error='Invalid password')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    write_audit(CONFIG['AUDIT_LOG_PATH'], 'logout', 'admin', {})
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT id, name FROM employees ORDER BY name')
    employees = c.fetchall()
    employee_data = []

    for emp in employees:
        days_used = calculate_eu_days(emp['id'])

        # Get total trip count
        c.execute('SELECT COUNT(*) FROM trips WHERE employee_id = ?', (emp['id'],))
        trip_count = c.fetchone()[0]

        # Get next upcoming trip
        today = datetime.now().date()
        c.execute('SELECT country, entry_date FROM trips WHERE employee_id = ? AND entry_date > ? ORDER BY entry_date ASC LIMIT 1',
                 (emp['id'], today.strftime('%Y-%m-%d')))
        next_trip = c.fetchone()

        # Get recent trips for the card view
        c.execute('SELECT country, entry_date, exit_date FROM trips WHERE employee_id = ? ORDER BY exit_date DESC LIMIT 5', (emp['id'],))
        recent_trips = c.fetchall()

        employee_data.append({
            'id': emp['id'],
            'name': emp['name'],
            'days_used': days_used,
            'days_remaining': 90 - days_used,
            'status': 'danger' if days_used > 80 else ('warning' if days_used > 60 else 'safe'),
            'trip_count': trip_count,
            'next_trip': next_trip,
            'recent_trips': recent_trips
        })

    conn.close()
    return render_template('dashboard.html', employees=employee_data)

@app.route('/employee/<int:employee_id>')
@login_required
def employee_detail(employee_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT name FROM employees WHERE id = ?', (employee_id,))
    employee = c.fetchone()
    if not employee:
        conn.close()
        return redirect(url_for('dashboard'))
    c.execute('SELECT id, country, entry_date, exit_date, travel_days FROM trips WHERE employee_id = ? ORDER BY entry_date DESC', (employee_id,))
    trips = c.fetchall()
    days_used = calculate_eu_days(employee_id)

    # Calculate days_until_safe if days_used >= 90
    days_until_safe = None
    if days_used >= 90:
        today = datetime.now().date()
        start_date = today - timedelta(days=180)
        # Find the oldest trip within the last 180 days
        c.execute('SELECT exit_date FROM trips WHERE employee_id = ? AND exit_date >= ? ORDER BY exit_date ASC LIMIT 1', (employee_id, start_date.strftime('%Y-%m-%d')))
        oldest_trip = c.fetchone()
        if oldest_trip:
            oldest_exit_date = datetime.strptime(oldest_trip['exit_date'], '%Y-%m-%d').date()
            # Calculate days until that trip falls out of the 180-day window
            # Trip falls out 181 days after exit date
            fall_out_date = oldest_exit_date + timedelta(days=181)
            days_until_safe = (fall_out_date - today).days

    # Process trips to add calculated fields
    processed_trips = []
    today = datetime.now().date()
    
    for trip in trips:
        entry_date = datetime.strptime(trip['entry_date'], '%Y-%m-%d').date()
        exit_date = datetime.strptime(trip['exit_date'], '%Y-%m-%d').date()
        
        # Calculate duration
        duration = (exit_date - entry_date).days + 1
        
        # Determine status
        if exit_date < today:
            status = 'completed'
        elif entry_date <= today <= exit_date:
            status = 'current'
        else:
            status = 'future'
        
        processed_trip = dict(trip)
        processed_trip['duration'] = duration
        processed_trip['status'] = status
        processed_trip['entry_date_obj'] = entry_date
        processed_trip['exit_date_obj'] = exit_date
        processed_trips.append(processed_trip)

    conn.close()
    return render_template('employee_detail.html', 
                         employee={'id': employee_id, 'name': employee['name']}, 
                         trips=processed_trips, 
                         days_used=days_used, 
                         days_until_safe=days_until_safe,
                         today=today)

@app.route('/employee/<int:employee_id>/calendar')
@login_required
def employee_calendar(employee_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT name FROM employees WHERE id = ?', (employee_id,))
    employee = c.fetchone()
    if not employee:
        conn.close()
        return redirect(url_for('dashboard'))

    # Get all trips for the employee
    c.execute('SELECT id, country, entry_date, exit_date FROM trips WHERE employee_id = ? ORDER BY entry_date DESC', (employee_id,))
    trips = c.fetchall()

    # Calculate current 180-day window
    today = datetime.now().date()
    window_start = today - timedelta(days=180)
    window_end = today

    # Calculate days used within the 180-day window
    days_used = calculate_eu_days(employee_id)

    # Prepare trip data with additional info for calendar display
    calendar_trips = []
    for trip in trips:
        entry_date = datetime.strptime(trip['entry_date'], '%Y-%m-%d').date()
        exit_date = datetime.strptime(trip['exit_date'], '%Y-%m-%d').date()

        # Check if trip overlaps with 180-day window
        overlaps_window = exit_date >= window_start and entry_date <= window_end

        # Calculate days within window for this trip
        if overlaps_window:
            effective_entry = max(entry_date, window_start)
            effective_exit = min(exit_date, window_end)
            days_in_window = (effective_exit - effective_entry).days + 1
        else:
            days_in_window = 0

        calendar_trips.append({
            'id': trip['id'],
            'country': trip['country'],
            'entry_date': entry_date,
            'exit_date': exit_date,
            'overlaps_window': overlaps_window,
            'days_in_window': days_in_window,
            'trip_duration': (exit_date - entry_date).days + 1
        })

    # Calculate timeline range (last 12 months)
    timeline_start = today - timedelta(days=365)
    timeline_end = today + timedelta(days=30)  # Show a bit into the future

    conn.close()
    return render_template('employee_calendar.html',
                         employee={'id': employee_id, 'name': employee['name']},
                         trips=calendar_trips,
                         days_used=days_used,
                         window_start=window_start,
                         window_end=window_end,
                         timeline_start=timeline_start,
                         timeline_end=timeline_end,
                         today=today)

@app.route('/calendar/<int:employee_id>')
@login_required
def calendar(employee_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT name FROM employees WHERE id = ?', (employee_id,))
    employee = c.fetchone()
    if not employee:
        conn.close()
        return redirect(url_for('dashboard'))

    # Calculate current 180-day window
    today = datetime.now().date()
    window_start = today - timedelta(days=180)
    window_end = today

    # Get trips for the specified date range (6 months ago to 3 months future)
    timeline_start = today - timedelta(days=180)  # 6 months ago
    timeline_end = today + timedelta(days=90)     # 3 months future

    c.execute('SELECT id, country, entry_date, exit_date FROM trips WHERE employee_id = ? AND exit_date >= ? AND entry_date <= ? ORDER BY entry_date',
              (employee_id, timeline_start.strftime('%Y-%m-%d'), timeline_end.strftime('%Y-%m-%d')))
    trips = c.fetchall()

    # Calculate days used within the 180-day window
    days_used = calculate_eu_days(employee_id)

    # Prepare trip data with color coding
    calendar_trips = []
    for trip in trips:
        entry_date = datetime.strptime(trip['entry_date'], '%Y-%m-%d').date()
        exit_date = datetime.strptime(trip['exit_date'], '%Y-%m-%d').date()

        # Determine trip type for color coding
        if exit_date < today:
            trip_type = 'past'
        elif entry_date <= today <= exit_date:
            trip_type = 'current'
        else:
            trip_type = 'future'

        # Check if trip overlaps with 180-day window
        overlaps_window = exit_date >= window_start and entry_date <= window_end

        # Calculate days within window for this trip
        if overlaps_window:
            effective_entry = max(entry_date, window_start)
            effective_exit = min(exit_date, window_end)
            days_in_window = (effective_exit - effective_entry).days + 1
        else:
            days_in_window = 0

        calendar_trips.append({
            'id': trip['id'],
            'country': trip['country'],
            'entry_date': entry_date,
            'exit_date': exit_date,
            'trip_type': trip_type,
            'overlaps_window': overlaps_window,
            'days_in_window': days_in_window,
            'trip_duration': (exit_date - entry_date).days + 1
        })

    conn.close()
    return render_template('calendar.html',
                         employee={'id': employee_id, 'name': employee['name']},
                         trips=calendar_trips,
                         days_used=days_used,
                         window_start=window_start,
                         window_end=window_end,
                         timeline_start=timeline_start,
                         timeline_end=timeline_end,
                         today=today)

def calculate_forecast_eu_days(employee_id, forecast_entry_date, forecast_exit_date):
    """Calculate EU days including a forecast trip without saving to database"""
    conn = get_db()
    c = conn.cursor()
    today = datetime.now().date()
    start_date = today - timedelta(days=180)

    # Get existing trips within the 180-day window
    c.execute('SELECT entry_date, exit_date FROM trips WHERE employee_id = ? AND exit_date >= ?',
              (employee_id, start_date.strftime('%Y-%m-%d')))
    existing_trips = c.fetchall()
    conn.close()

    total_days = 0

    # Calculate days for existing trips
    for trip in existing_trips:
        entry = datetime.strptime(trip['entry_date'], '%Y-%m-%d').date()
        exit = datetime.strptime(trip['exit_date'], '%Y-%m-%d').date()
        effective_entry = max(entry, start_date)
        effective_exit = min(exit, today)
        if effective_entry <= effective_exit:
            total_days += (effective_exit - effective_entry).days + 1

    # Calculate days for forecast trip that would be within 180-day window during the trip period
    forecast_entry = datetime.strptime(forecast_entry_date, '%Y-%m-%d').date()
    forecast_exit = datetime.strptime(forecast_exit_date, '%Y-%m-%d').date()

    # The 180-day window during the forecast trip would be from (forecast_exit - 180) to forecast_exit
    forecast_window_start = forecast_exit - timedelta(days=180)
    forecast_window_end = forecast_exit

    # Recalculate existing trips that would be within the window during the forecast trip
    forecast_total = 0
    for trip in existing_trips:
        entry = datetime.strptime(trip['entry_date'], '%Y-%m-%d').date()
        exit = datetime.strptime(trip['exit_date'], '%Y-%m-%d').date()
        effective_entry = max(entry, forecast_window_start)
        effective_exit = min(exit, forecast_window_end)
        if effective_entry <= effective_exit:
            forecast_total += (effective_exit - effective_entry).days + 1

    # Add forecast trip days that fall within its own 180-day window
    forecast_effective_entry = max(forecast_entry, forecast_window_start)
    forecast_effective_exit = min(forecast_exit, forecast_window_end)
    if forecast_effective_entry <= forecast_effective_exit:
        forecast_total += (forecast_effective_exit - forecast_effective_entry).days + 1

    return {
        'current_days_used': total_days,
        'forecast_total_days': forecast_total,
        'forecast_trip_days': (forecast_exit - forecast_entry).days + 1
    }

@app.route('/forecast-trip', methods=['POST'])
@login_required
def forecast_trip():
    try:
        data = request.get_json()
        employee_id = int(data.get('employee_id'))
        entry_date = data.get('entry_date')
        exit_date = data.get('exit_date')
        country = data.get('country', '')

        # Validate inputs
        if not all([employee_id, entry_date, exit_date]):
            return jsonify({
                'error': 'Missing required fields: employee_id, entry_date, exit_date'
            }), 400

        # Validate date format and logic
        try:
            entry_dt = datetime.strptime(entry_date, '%Y-%m-%d').date()
            exit_dt = datetime.strptime(exit_date, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

        if exit_dt < entry_dt:
            return jsonify({'error': 'Exit date cannot be before entry date'}), 400

        # Check if employee exists
        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT name FROM employees WHERE id = ?', (employee_id,))
        employee = c.fetchone()
        conn.close()

        if not employee:
            return jsonify({'error': 'Employee not found'}), 404

        # Calculate forecast
        current_days_used = calculate_eu_days(employee_id)
        forecast_data = calculate_forecast_eu_days(employee_id, entry_date, exit_date)

        total_days_after = forecast_data['forecast_total_days']
        trip_days = forecast_data['forecast_trip_days']

        # Determine status
        if total_days_after > 90:
            status = 'over_limit'
            message = f'❌ This trip would exceed the 90-day limit by {total_days_after - 90} days.'
        elif total_days_after >= 85:
            status = 'warning'
            message = f'⚠️ This trip brings you close to the limit. Only {90 - total_days_after} days remaining.'
        else:
            status = 'valid'
            message = f'✅ This trip is valid. You\'ll have {90 - total_days_after} days remaining.'

        return jsonify({
            'employee_name': employee['name'],
            'current_days_used': current_days_used,
            'planned_trip_days': trip_days,
            'total_days_after': total_days_after,
            'status': status,
            'message': message,
            'country': country,
            'entry_date': entry_date,
            'exit_date': exit_date,
            'remaining_days': 90 - total_days_after
        })

    except Exception as e:
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@app.route('/add_employee', methods=['POST'])
@login_required
def add_employee():
    name = re.sub(r'[<>"\'&]', '', request.form.get('name', '').strip())
    if name:
        conn = get_db()
        try:
            conn.execute('INSERT INTO employees (name) VALUES (?)', (name,))
            conn.commit()
        except:
            pass
        conn.close()
    return redirect(url_for('dashboard'))

@app.route('/add_trip', methods=['POST'])
@login_required
def add_trip():
    employee_id = request.form.get('employee_id')
    country_code = (request.form.get('country_code', '') or '').strip().upper()
    country_code = re.sub(r'[^A-Z]', '', country_code)
    entry_date = request.form.get('entry_date', '')
    exit_date = request.form.get('exit_date', '')
    try:
        if datetime.strptime(exit_date, '%Y-%m-%d') >= datetime.strptime(entry_date, '%Y-%m-%d'):
            if country_code and COUNTRY_CODE_MAPPING.get(country_code):
                conn = get_db()
                conn.execute('INSERT INTO trips (employee_id, country, entry_date, exit_date) VALUES (?, ?, ?, ?)', (employee_id, country_code, entry_date, exit_date))
                conn.commit()
                conn.close()
    except:
        pass
    return redirect(url_for('employee_detail', employee_id=employee_id))

@app.route('/delete_trip/<int:trip_id>', methods=['POST'])
@login_required
def delete_trip(trip_id):
    employee_id = request.form.get('employee_id')
    conn = get_db()
    conn.execute('DELETE FROM trips WHERE id = ?', (trip_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('employee_detail', employee_id=employee_id))

@app.route('/delete_employee/<int:employee_id>', methods=['POST'])
@login_required
def delete_employee(employee_id):
    conn = get_db()
    conn.execute('DELETE FROM employees WHERE id = ?', (employee_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('dashboard'))

@app.route('/change_password', methods=['POST'])
@login_required
def change_password():
    new_password = request.form.get('new_password', '')
    if len(new_password) >= 6:
        conn = get_db()
        conn.execute('UPDATE admin SET password_hash = ? WHERE id = 1', (hasher.hash(new_password),))
        conn.commit()
        conn.close()
    return redirect(url_for('dashboard'))

@app.route('/bulk_add_trip', methods=['GET', 'POST'])
@login_required
def bulk_add_trip():
    if request.method == 'POST':
        employee_ids = request.form.getlist('employee_ids')
        country_code = (request.form.get('country_code', '') or '').strip().upper()
        country_code = re.sub(r'[^A-Z]', '', country_code)
        entry_date = request.form.get('entry_date', '')
        exit_date = request.form.get('exit_date', '')

        # Validation
        if not employee_ids:
            conn = get_db()
            c = conn.cursor()
            c.execute('SELECT id, name FROM employees ORDER BY name')
            employees = c.fetchall()
            conn.close()
            return render_template('bulk_add_trip.html', employees=employees, error='Please select at least one employee.')

        if not country_code or not entry_date or not exit_date:
            conn = get_db()
            c = conn.cursor()
            c.execute('SELECT id, name FROM employees ORDER BY name')
            employees = c.fetchall()
            conn.close()
            return render_template('bulk_add_trip.html', employees=employees, error='Please fill in all fields.')

        try:
            if datetime.strptime(exit_date, '%Y-%m-%d') < datetime.strptime(entry_date, '%Y-%m-%d'):
                conn = get_db()
                c = conn.cursor()
                c.execute('SELECT id, name FROM employees ORDER BY name')
                employees = c.fetchall()
                conn.close()
                return render_template('bulk_add_trip.html', employees=employees, error='Exit date cannot be before entry date.')
        except ValueError:
            conn = get_db()
            c = conn.cursor()
            c.execute('SELECT id, name FROM employees ORDER BY name')
            employees = c.fetchall()
            conn.close()
            return render_template('bulk_add_trip.html', employees=employees, error='Invalid date format.')

        # Insert trips for selected employees
        conn = get_db()
        success_count = 0
        for employee_id in employee_ids:
            try:
                if country_code and COUNTRY_CODE_MAPPING.get(country_code):
                    conn.execute('INSERT INTO trips (employee_id, country, entry_date, exit_date) VALUES (?, ?, ?, ?)',
                               (employee_id, country_code, entry_date, exit_date))
                    success_count += 1
                else:
                    continue
            except:
                continue

        conn.commit()
        conn.close()

        if success_count > 0:
            message = f'Successfully added trip to {success_count} employee{"s" if success_count != 1 else ""}.'
            return redirect(url_for('dashboard'))
        else:
            conn = get_db()
            c = conn.cursor()
            c.execute('SELECT id, name FROM employees ORDER BY name')
            employees = c.fetchall()
            conn.close()
            return render_template('bulk_add_trip.html', employees=employees, error='Failed to add trips. Please try again.')

    # GET request - show the form
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT id, name FROM employees ORDER BY name')
    employees = c.fetchall()
    conn.close()
    return render_template('bulk_add_trip.html', employees=employees)

@app.route('/import_excel', methods=['GET', 'POST'])
@login_required
def import_excel():
    if request.method == 'POST':
        # Handle file upload
        if 'excel_file' not in request.files:
            flash('No file uploaded')
            return redirect(request.url)

        file = request.files['excel_file']
        if file.filename == '':
            flash('No file selected')
            return redirect(request.url)

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)

            try:
                # Use importer.process_excel to parse and save, returning a summary
                from importer import process_excel
                summary = process_excel(filepath, db_path=DATABASE)

                # Clean up uploaded file
                os.remove(filepath)

                # Flash basic message
                if summary.get('trips_added', 0) > 0:
                    flash(f"Imported {summary['trips_added']} trips. {summary.get('duplicates_skipped', 0)} duplicates skipped. {summary.get('uk_jobs_skipped', 0)} UK jobs ignored.")
                else:
                    flash('No valid trips found in the uploaded file.')

                # Render page with detailed summary
                return render_template('import_excel.html', import_summary=summary)

            except Exception as e:
                # Clean up uploaded file
                os.remove(filepath)
                flash(f'Error importing Excel file: {str(e)}')
                print(f"Import error: {e}")
                return redirect(request.url)
        else:
            flash('Invalid file format. Please upload an Excel file (.xlsx or .xls)')
            return redirect(request.url)

    return render_template('import_excel.html')

@app.route('/import_preview')
@login_required
def import_preview():
    if 'import_preview' not in session:
        flash('No import data found')
        return redirect(url_for('import_excel'))

    data = session['import_preview']
    return render_template('import_preview.html',
                         valid_trips=data['valid_trips'],
                         errors=data['errors'])

@app.route('/confirm_import', methods=['POST'])
@login_required
def confirm_import():
    if 'import_preview' not in session:
        flash('No import data found')
        return redirect(url_for('import_excel'))

    valid_trips = session['import_preview']['valid_trips']

    # Import the trips
    conn = get_db()
    c = conn.cursor()
    success_count = 0

    for trip in valid_trips:
        try:
            c.execute('INSERT INTO trips (employee_id, country, entry_date, exit_date) VALUES (?, ?, ?, ?)',
                     (trip['employee_id'], trip['country'], trip['entry_date'], trip['exit_date']))
            success_count += 1
        except Exception as e:
            print(f"Error importing trip: {e}")

    conn.commit()
    conn.close()

    # Clear session data
    session.pop('import_preview', None)

    flash(f'Successfully imported {success_count} trips')
    return redirect(url_for('dashboard'))

@app.route('/cancel_import', methods=['POST'])
@login_required
def cancel_import():
    session.pop('import_preview', None)
    flash('Import cancelled')
    return redirect(url_for('import_excel'))

# Testing utility: clear all trips
@app.route('/clear_imported_trips', methods=['POST'])
@login_required
def clear_imported_trips():
    try:
        conn = get_db()
        c = conn.cursor()
        c.execute('DELETE FROM trips')
        conn.commit()
        conn.close()
        flash('All trips have been cleared for testing.')
    except Exception as e:
        flash(f'Failed to clear trips: {str(e)}')
    return redirect(url_for('import_excel'))

# -------------------------
# DSAR (Data Subject Access Request) endpoints
# -------------------------

@app.route('/api/dsar/export')
@login_required
def dsar_export():
    """Export personal data for a given employee (by name or id)."""
    employee_name = request.args.get('employee_name', '').strip()
    employee_id = request.args.get('employee_id', '').strip()

    conn = get_db()
    c = conn.cursor()

    if employee_id:
        c.execute('SELECT id, name FROM employees WHERE id = ?', (employee_id,))
    elif employee_name:
        c.execute('SELECT id, name FROM employees WHERE LOWER(name) = LOWER(?)', (employee_name,))
    else:
        conn.close()
        return jsonify({'error': 'Provide employee_name or employee_id'}), 400

    emp = c.fetchone()
    if not emp:
        conn.close()
        return jsonify({'error': 'Employee not found'}), 404

    c.execute('SELECT country, entry_date, exit_date, travel_days FROM trips WHERE employee_id = ? ORDER BY entry_date ASC', (emp[0],))
    trips = [
        {
            'country': row[0],
            'entry_date': row[1],
            'exit_date': row[2],
            'travel_days': row[3] if row[3] is not None else 0
        }
        for row in c.fetchall()
    ]
    conn.close()

    export_payload = {
        'employee': {
            'id': emp[0],
            'name': emp[1]
        },
        'trips': trips,
        'generated_at': datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')
    }

    response = make_response(jsonify(export_payload))
    response.headers['Content-Disposition'] = f'attachment; filename=eu-trip-tracker-{emp[1]}-export.json'
    return response


@app.route('/api/dsar/delete', methods=['POST'])
@login_required
def dsar_delete():
    """Delete/anonymize personal data for a given employee (by id or name)."""
    try:
        data = request.get_json() or {}
    except Exception:
        data = {}

    employee_id = str(data.get('employee_id', '')).strip()
    employee_name = str(data.get('employee_name', '')).strip()

    conn = get_db()
    c = conn.cursor()

    # Resolve employee
    if employee_id:
        c.execute('SELECT id, name FROM employees WHERE id = ?', (employee_id,))
    elif employee_name:
        c.execute('SELECT id, name FROM employees WHERE LOWER(name) = LOWER(?)', (employee_name,))
    else:
        conn.close()
        return jsonify({'error': 'Provide employee_name or employee_id'}), 400

    emp = c.fetchone()
    if not emp:
        conn.close()
        return jsonify({'error': 'Employee not found'}), 404

    # Delete trips and employee record
    c.execute('DELETE FROM trips WHERE employee_id = ?', (emp[0],))
    c.execute('DELETE FROM employees WHERE id = ?', (emp[0],))
    conn.commit()
    conn.close()

    return jsonify({'status': 'deleted', 'employee_id': emp[0], 'employee_name': emp[1]})

@app.route('/calendar_view')
@login_required
def calendar_view():
    # Render global calendar view with employee filter
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT id, name FROM employees ORDER BY name')
    employees = c.fetchall()
    conn.close()
    today = datetime.now().date()
    return render_template('global_calendar.html', employees=employees, today=today)

@app.route('/api/calendar_data')
@login_required
def api_calendar_data():
    try:
        # Parse optional params from FullCalendar
        start_param = request.args.get('start')
        end_param = request.args.get('end')
        employee_id_param = request.args.get('employee_id')

        today = datetime.now().date()
        allowed_start = today - timedelta(days=180)
        allowed_end = today + timedelta(days=56)

        # FullCalendar passes ISO strings (YYYY-MM-DD)
        def parse_date_or_default(value, default_value):
            try:
                if value:
                    return datetime.strptime(value[:10], '%Y-%m-%d').date()
            except Exception:
                pass
            return default_value

        requested_start = parse_date_or_default(start_param, allowed_start)
        requested_end = parse_date_or_default(end_param, allowed_end)

        # Clamp to allowed window
        query_start = max(requested_start, allowed_start)
        query_end = min(requested_end, allowed_end)

        conn = get_db()
        conn.row_factory = sqlite3.Row
        c = conn.cursor()

        # For accurate 180-day calculations near the start boundary, prefetch from (allowed_start - 180)
        prefetch_start = allowed_start - timedelta(days=180)

        sql = (
            'SELECT trips.id as trip_id, trips.employee_id, employees.name as employee_name, '
            'trips.country, trips.entry_date, trips.exit_date '
            'FROM trips JOIN employees ON trips.employee_id = employees.id '
            'WHERE trips.exit_date >= ? AND trips.entry_date <= ?'
        )
        params = [query_start.strftime('%Y-%m-%d'), query_end.strftime('%Y-%m-%d')]

        if employee_id_param:
            sql += ' AND trips.employee_id = ?'
            params.append(employee_id_param)

        sql += ' ORDER BY trips.entry_date ASC'

        c.execute(sql, params)
        rows = c.fetchall()

        # Prefetch trips for 180-day calculations per employee
        employee_ids = set([row['employee_id'] for row in rows])
        trips_by_employee = {}
        if employee_ids:
            if employee_id_param:
                prefetch_ids = [int(employee_id_param)]
            else:
                prefetch_ids = list(employee_ids)
            placeholders = ','.join(['?'] * len(prefetch_ids))
            pre_sql = (
                f'SELECT employee_id, entry_date, exit_date FROM trips '
                f'WHERE employee_id IN ({placeholders}) AND exit_date >= ? AND entry_date <= ?'
            )
            pre_params = prefetch_ids + [prefetch_start.strftime('%Y-%m-%d'), allowed_end.strftime('%Y-%m-%d')]
            c.execute(pre_sql, pre_params)
            pre_rows = c.fetchall()
            for r in pre_rows:
                trips_by_employee.setdefault(r['employee_id'], []).append({'entry_date': r['entry_date'], 'exit_date': r['exit_date']})

        events = []
        GREEN = '#10b981'
        YELLOW = '#f59e0b'
        RED = '#ef4444'

        for row in rows:
            emp_id = row['employee_id']
            emp_name = row['employee_name']
            country = row['country']
            entry_dt = datetime.strptime(row['entry_date'], '%Y-%m-%d').date()
            exit_dt = datetime.strptime(row['exit_date'], '%Y-%m-%d').date()

            # Duration inclusive
            duration_days = (exit_dt - entry_dt).days + 1

            # Current usage (as of today)
            emp_trips = trips_by_employee.get(emp_id, [])
            current_used = calculate_eu_days_at_date(emp_trips, today)

            # Forecast usage at the end of this trip
            forecast_used = calculate_eu_days_at_date(emp_trips, exit_dt)

            # Color rules
            if forecast_used >= 90:
                color = RED
                status = 'over_limit'
            elif forecast_used >= 80:
                color = YELLOW
                status = 'warning'
            else:
                color = GREEN
                status = 'ok'

            # FullCalendar's end is exclusive; add one day
            end_exclusive = (exit_dt + timedelta(days=1)).strftime('%Y-%m-%d')

            tooltip = (
                f"Employee: {emp_name}\n"
                f"Country: {country}\n"
                f"Trip: {row['entry_date']} → {row['exit_date']}\n"
                f"Duration: {duration_days} days\n"
                f"Used now: {current_used}/90\n"
                f"Used at trip end: {forecast_used}/90"
            )

            events.append({
                'id': f"{row['trip_id']}",
                'title': f"{emp_name} - {country}",
                'start': row['entry_date'],
                'end': end_exclusive,
                'allDay': True,
                'color': color,
                'extendedProps': {
                    'employeeId': emp_id,
                    'employeeName': emp_name,
                    'country': country,
                    'durationDays': duration_days,
                    'currentUsed': current_used,
                    'forecastUsed': forecast_used,
                    'status': status,
                    'tooltip': tooltip
                }
            })

        conn.close()
        return jsonify(events)
    except Exception as e:
        return jsonify({'error': f'Server error: {str(e)}'}), 500

# -------------------------
# GDPR Compliance Endpoints
# -------------------------

@app.route('/admin/privacy-tools')
@login_required
def admin_privacy_tools():
    """Privacy tools page for DSAR operations."""
    try:
        expired_trips = get_expired_trips(DATABASE, CONFIG.get('RETENTION_MONTHS', 36))
        
        # Get last purge timestamp from a simple file
        last_purge_file = os.path.join(CONFIG.get('DSAR_EXPORT_DIR', './exports'), '.last_purge')
        last_purge = None
        if os.path.exists(last_purge_file):
            try:
                with open(last_purge_file, 'r') as f:
                    last_purge = f.read().strip()
            except:
                pass
        
        return render_template('admin_privacy_tools.html',
                             retention_months=CONFIG.get('RETENTION_MONTHS', 36),
                             expired_count=len(expired_trips),
                             last_purge=last_purge)
    except Exception as e:
        flash(f'Error loading privacy tools: {str(e)}')
        return redirect(url_for('dashboard'))

@app.route('/admin/settings', methods=['GET', 'POST'])
@login_required
def admin_settings():
    """Admin settings page."""
    if request.method == 'POST':
        try:
            updates = {
                'RETENTION_MONTHS': int(request.form.get('retention_months', 36)),
                'SESSION_IDLE_TIMEOUT_MINUTES': int(request.form.get('session_timeout', 30)),
                'PASSWORD_HASH_SCHEME': request.form.get('password_scheme', 'argon2'),
                'DSAR_EXPORT_DIR': request.form.get('export_dir', './exports'),
                'AUDIT_LOG_PATH': request.form.get('audit_log', './logs/audit.log'),
                'SESSION_COOKIE_SECURE': 'session_cookie_secure' in request.form
            }
            
            global CONFIG
            CONFIG = save_config(updates)
            
            write_audit(CONFIG['AUDIT_LOG_PATH'], 'settings_updated', 'admin', updates)
            flash('Settings updated successfully')
            return redirect(url_for('admin_settings'))
        except Exception as e:
            flash(f'Error updating settings: {str(e)}')
    
    return render_template('admin_settings.html', config=CONFIG, message=request.args.get('message'), error=request.args.get('error'))

@app.route('/api/employees/search')
@login_required
def api_employees_search():
    """Search employees by name."""
    try:
        query = request.args.get('q', '').strip()
        if not query:
            return jsonify({'error': 'No search query provided'}), 400
        
        conn = get_db()
        c = conn.cursor()
        c.execute('''
            SELECT e.id, e.name, COUNT(t.id) as trip_count
            FROM employees e
            LEFT JOIN trips t ON e.id = t.employee_id
            WHERE LOWER(e.name) LIKE LOWER(?)
            GROUP BY e.id, e.name
            ORDER BY e.name
            LIMIT 10
        ''', (f'%{query}%',))
        
        employees = []
        for row in c.fetchall():
            employees.append({
                'id': row['id'],
                'name': row['name'],
                'trip_count': row['trip_count']
            })
        
        conn.close()
        return jsonify({'employees': employees})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/admin/dsar/export/<int:employee_id>')
@login_required
def admin_dsar_export(employee_id):
    """Generate and download DSAR export ZIP for employee."""
    try:
        result = create_dsar_export(
            DATABASE, 
            employee_id, 
            CONFIG.get('DSAR_EXPORT_DIR', './exports'),
            CONFIG.get('RETENTION_MONTHS', 36)
        )
        
        if not result.get('success'):
            flash(result.get('error', 'Unknown error'))
            return redirect(url_for('admin_privacy_tools'))
        
        write_audit(CONFIG['AUDIT_LOG_PATH'], 'dsar_export', 'admin', {
            'employee_id': employee_id,
            'employee_name': result['employee_name']
        })
        
        return send_file(
            result['file_path'],
            as_attachment=True,
            download_name=result['filename'],
            mimetype='application/zip'
        )
    except Exception as e:
        flash(f'Error exporting data: {str(e)}')
        return redirect(url_for('admin_privacy_tools'))

@app.route('/admin/dsar/delete/<int:employee_id>', methods=['POST'])
@login_required
def admin_dsar_delete(employee_id):
    """Delete employee and all associated data."""
    try:
        result = delete_employee_data(DATABASE, employee_id)
        
        if not result.get('success'):
            return jsonify(result), 404
        
        write_audit(CONFIG['AUDIT_LOG_PATH'], 'dsar_delete', 'admin', {
            'employee_id': employee_id,
            'employee_name': result['employee_name'],
            'trips_deleted': result['trips_deleted']
        })
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/dsar/rectify/<int:employee_id>', methods=['POST'])
@login_required
def admin_dsar_rectify(employee_id):
    """Rectify (update) employee name."""
    try:
        data = request.get_json() or {}
        new_name = data.get('new_name', '').strip()
        
        result = rectify_employee_name(DATABASE, employee_id, new_name)
        
        if not result.get('success'):
            return jsonify(result), 400
        
        write_audit(CONFIG['AUDIT_LOG_PATH'], 'dsar_rectify', 'admin', {
            'employee_id': employee_id,
            'old_name': result['old_name'],
            'new_name': result['new_name']
        })
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/dsar/anonymize/<int:employee_id>', methods=['POST'])
@login_required
def admin_dsar_anonymize(employee_id):
    """Anonymize employee name."""
    try:
        result = anonymize_employee(DATABASE, employee_id)
        
        if not result.get('success'):
            return jsonify(result), 404
        
        write_audit(CONFIG['AUDIT_LOG_PATH'], 'dsar_anonymize', 'admin', {
            'employee_id': employee_id,
            'old_name': result['old_name'],
            'new_name': result['new_name']
        })
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/retention/purge', methods=['POST'])
@login_required
def admin_retention_purge():
    """Run retention purge to delete old trips."""
    try:
        result = purge_expired_trips(DATABASE, CONFIG.get('RETENTION_MONTHS', 36))
        
        # Save last purge timestamp
        last_purge_file = os.path.join(CONFIG.get('DSAR_EXPORT_DIR', './exports'), '.last_purge')
        os.makedirs(os.path.dirname(last_purge_file), exist_ok=True)
        with open(last_purge_file, 'w') as f:
            f.write(datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC'))
        
        write_audit(CONFIG['AUDIT_LOG_PATH'], 'retention_purge', 'admin', result)
        
        return jsonify({
            'success': True,
            'trips_deleted': result['trips_deleted'],
            'employees_deleted': result['employees_deleted'],
            'cutoff_date': result['cutoff_date']
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/retention/preview')
@login_required
def api_retention_preview():
    """Preview trips that would be purged."""
    try:
        expired_trips = get_expired_trips(DATABASE, CONFIG.get('RETENTION_MONTHS', 36))
        
        # Count unique employees
        employee_ids = set([trip['employee_id'] for trip in expired_trips])
        
        return jsonify({
            'trips_count': len(expired_trips),
            'employees_affected': len(employee_ids)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/admin/retention/expired')
@login_required
def admin_retention_expired():
    """Show list of expired trips."""
    try:
        expired_trips = get_expired_trips(DATABASE, CONFIG.get('RETENTION_MONTHS', 36))
        return render_template('expired_trips.html', 
                             trips=expired_trips,
                             retention_months=CONFIG.get('RETENTION_MONTHS', 36))
    except Exception as e:
        flash(f'Error loading expired trips: {str(e)}')
        return redirect(url_for('admin_privacy_tools'))

@app.route('/export/trips/csv')
@login_required
def export_trips_csv_endpoint():
    """Export all trips as CSV."""
    try:
        employee_id = request.args.get('employee_id', type=int)
        csv_data = export_trips_csv(DATABASE, employee_id)
        
        response = make_response(csv_data)
        filename = f"trips_employee_{employee_id}.csv" if employee_id else "trips_all.csv"
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'text/csv'
        
        write_audit(CONFIG['AUDIT_LOG_PATH'], 'export_csv', 'admin', {'employee_id': employee_id})
        return response
    except Exception as e:
        flash(f'Error exporting CSV: {str(e)}')
        return redirect(url_for('dashboard'))

@app.route('/export/employee/<int:employee_id>/pdf')
@login_required
def export_employee_pdf(employee_id):
    """Export employee report as PDF."""
    try:
        pdf_bytes = export_employee_report_pdf(DATABASE, employee_id)
        
        # Get employee name for filename
        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT name FROM employees WHERE id = ?', (employee_id,))
        emp = c.fetchone()
        conn.close()
        
        filename = f"report_{emp['name'].replace(' ', '_')}.pdf" if emp else f"report_{employee_id}.pdf"
        
        response = make_response(pdf_bytes)
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'application/pdf'
        
        write_audit(CONFIG['AUDIT_LOG_PATH'], 'export_pdf', 'admin', {'employee_id': employee_id})
        return response
    except Exception as e:
        flash(f'Error exporting PDF: {str(e)}')
        return redirect(url_for('employee_detail', employee_id=employee_id))

@app.route('/export/all/pdf')
@login_required
def export_all_pdf():
    """Export all employees summary as PDF."""
    try:
        pdf_bytes = export_all_employees_report_pdf(DATABASE)
        
        response = make_response(pdf_bytes)
        response.headers['Content-Disposition'] = 'attachment; filename=compliance_summary.pdf'
        response.headers['Content-Type'] = 'application/pdf'
        
        write_audit(CONFIG['AUDIT_LOG_PATH'], 'export_all_pdf', 'admin', {})
        return response
    except Exception as e:
        flash(f'Error exporting PDF: {str(e)}')
        return redirect(url_for('dashboard'))

if __name__ == '__main__':
    init_db()
    print("\n" + "="*50)
    print("EU TRIP TRACKER")
    print("="*50)
    print("Password: admin123")
    print("URL: http://127.0.0.1:5001")
    print("="*50 + "\n")
    app.run(debug=True, host='127.0.0.1', port=5001)
