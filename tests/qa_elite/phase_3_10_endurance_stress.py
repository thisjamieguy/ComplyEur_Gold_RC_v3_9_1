"""
Elite QA Task Force: Phase 3.10 Endurance & Stress Testing
180-minute continuous stress test of ComplyEur's calculation engine, backend integrity, and frontend synchronization.

Mission: Prove ComplyEur remains mathematically flawless and stable under continuous stress.
"""

import os
import sys
import time
import json
import sqlite3
import tempfile
import threading
import subprocess
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
import re
from collections import defaultdict
import traceback

# Add project root to path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../..')))

import openpyxl
import pandas as pd
import psutil

try:
    from playwright.sync_api import sync_playwright, Page, Browser, BrowserContext
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False
    Page = Browser = BrowserContext = None

try:
    from diff_match_patch import diff_match_patch
    DIFF_LIB_AVAILABLE = True
except ImportError:
    DIFF_LIB_AVAILABLE = False

from importer import process_excel_file
from app.services.rolling90 import presence_days, days_used_in_window, calculate_days_remaining
from app.services.exports import export_trips_csv, export_all_employees_report_pdf, calculate_eu_days_from_trips

# Test configuration
TEST_DURATION_MINUTES = int(os.getenv('TEST_DURATION_MINUTES', '180'))  # 3 hours default
TEST_DURATION_SECONDS = TEST_DURATION_MINUTES * 60

# Report directories
REPORTS_DIR = Path(__file__).parent.parent.parent / 'reports' / 'qa_elite' / 'endurance'
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

# Tolerance for floating-point drift
DRIFT_TOLERANCE = 0.0000001

# Frontend synchronization lag threshold (milliseconds)
SYNC_LAG_THRESHOLD_MS = 50


class SystemMonitor:
    """Monitor CPU, memory, and I/O performance"""
    
    def __init__(self):
        self.metrics = []
        self.process = psutil.Process()
        self.start_time = time.time()
    
    def record_metrics(self):
        """Record current system metrics"""
        # Get I/O counters if available (may not be available on all platforms)
        try:
            io_counters = self.process.io_counters()
            io_read_mb = io_counters.read_bytes / 1024 / 1024
            io_write_mb = io_counters.write_bytes / 1024 / 1024
        except (AttributeError, OSError, NotImplementedError):
            # I/O counters not available on this platform or process type
            io_read_mb = 0.0
            io_write_mb = 0.0
        
        metric = {
            'timestamp': datetime.now().isoformat(),
            'elapsed_seconds': time.time() - self.start_time,
            'cpu_percent': self.process.cpu_percent(interval=0.1),
            'memory_mb': self.process.memory_info().rss / 1024 / 1024,
            'memory_percent': self.process.memory_percent(),
            'io_read_mb': io_read_mb,
            'io_write_mb': io_write_mb,
        }
        self.metrics.append(metric)
        return metric
    
    def get_metrics(self) -> List[Dict]:
        """Get all recorded metrics"""
        return self.metrics
    
    def get_summary(self) -> Dict:
        """Get summary statistics"""
        if not self.metrics:
            return {}
        
        cpu_values = [m['cpu_percent'] for m in self.metrics]
        memory_values = [m['memory_mb'] for m in self.metrics]
        
        return {
            'duration_seconds': self.metrics[-1]['elapsed_seconds'],
            'cpu_avg': sum(cpu_values) / len(cpu_values),
            'cpu_max': max(cpu_values),
            'cpu_min': min(cpu_values),
            'memory_avg_mb': sum(memory_values) / len(memory_values),
            'memory_max_mb': max(memory_values),
            'memory_min_mb': min(memory_values),
            'memory_growth_mb': memory_values[-1] - memory_values[0] if len(memory_values) > 1 else 0,
            'total_io_read_mb': self.metrics[-1]['io_read_mb'],
            'total_io_write_mb': self.metrics[-1]['io_write_mb'],
        }


class ExcelDatasetGenerator:
    """Generate Excel datasets for load testing"""
    
    SCHENGEN_COUNTRIES = ['AT', 'BE', 'BG', 'HR', 'CY', 'CZ', 'DK', 'EE', 'ES', 'FI', 'FR', 
                          'DE', 'GR', 'HU', 'IS', 'IE', 'IT', 'LV', 'LI', 'LT', 'LU', 'MT', 
                          'NL', 'NO', 'PL', 'PT', 'RO', 'SK', 'SI', 'SE', 'CH']
    
    @staticmethod
    def generate_excel_dataset(output_path: str, employee_count: int, 
                               start_date: date = None, end_date: date = None) -> str:
        """Generate Excel file with randomized employee travel data"""
        if start_date is None:
            start_date = date.today() - timedelta(days=180)
        if end_date is None:
            end_date = date.today() + timedelta(days=180)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employee Travel"
        
        # Header row
        header_row = 3
        ws.cell(row=header_row, column=1, value="Employee")
        
        # Generate date columns
        current_date = start_date
        col = 2
        date_columns = {}
        while current_date <= end_date:
            ws.cell(row=header_row, column=col, value=current_date.strftime('%d/%m/%Y'))
            date_columns[col] = current_date
            col += 1
            current_date += timedelta(days=1)
        
        # Generate employee rows with random travel data
        import random
        row = header_row + 1
        
        for emp_idx in range(employee_count):
            employee_name = f"Employee_{emp_idx+1:03d}"
            ws.cell(row=row, column=1, value=employee_name)
            
            # Generate 3-10 random trips per employee
            num_trips = random.randint(3, 10)
            trips = []
            
            for _ in range(num_trips):
                # Random trip dates
                trip_start_offset = random.randint(0, (end_date - start_date).days - 10)
                trip_start = start_date + timedelta(days=trip_start_offset)
                trip_duration = random.randint(2, 14)
                trip_end = min(trip_start + timedelta(days=trip_duration - 1), end_date)
                
                # Random country
                country = random.choice(ExcelDatasetGenerator.SCHENGEN_COUNTRIES)
                is_travel = random.choice([True, False])
                
                trips.append({
                    'start': trip_start,
                    'end': trip_end,
                    'country': country,
                    'is_travel': is_travel
                })
            
            # Fill cells for trips
            for trip in trips:
                start_col = None
                end_col = None
                
                # Find column indices
                for col_idx, col_date in date_columns.items():
                    if col_date == trip['start']:
                        start_col = col_idx
                    if col_date == trip['end']:
                        end_col = col_idx
                
                if start_col and end_col:
                    cell_text = trip['country']
                    if trip['is_travel']:
                        cell_text = f"tr/{cell_text}"
                    
                    for trip_col in range(start_col, end_col + 1):
                        ws.cell(row=row, column=trip_col, value=cell_text)
            
            row += 1
        
        wb.save(output_path)
        return output_path


class CalculationDriftTracker:
    """Track floating-point drift in calculations"""
    
    def __init__(self):
        self.drift_log = []
        self.baseline_values = {}
    
    def record_baseline(self, employee_name: str, days_used: float):
        """Record baseline calculation value"""
        self.baseline_values[employee_name] = days_used
    
    def check_drift(self, employee_name: str, days_used: float) -> Optional[Dict]:
        """Check for drift from baseline"""
        if employee_name not in self.baseline_values:
            self.baseline_values[employee_name] = days_used
            return None
        
        baseline = self.baseline_values[employee_name]
        drift = abs(days_used - baseline)
        
        if drift > DRIFT_TOLERANCE:
            drift_entry = {
                'timestamp': datetime.now().isoformat(),
                'employee': employee_name,
                'baseline': baseline,
                'current': days_used,
                'drift': drift
            }
            self.drift_log.append(drift_entry)
            return drift_entry
        
        return None
    
    def get_drift_log(self) -> List[Dict]:
        """Get all drift entries"""
        return self.drift_log


class EnduranceStressTestSuite:
    """Main endurance stress test suite"""
    
    def __init__(self, db_path: str, app_url: str = 'http://127.0.0.1:5001', 
                 admin_password: str = 'admin123', admin_username: str = 'admin'):
        self.db_path = db_path
        self.app_url = app_url
        self.admin_password = admin_password
        self.admin_username = admin_username
        
        self.start_time = None
        self.end_time = None
        
        # Test modules
        self.monitor = SystemMonitor()
        self.drift_tracker = CalculationDriftTracker()
        
        # Results tracking
        self.results = {
            'start_time': None,
            'end_time': None,
            'duration_seconds': 0,
            'excel_import_tests': {
                'runs': 0,
                'successes': 0,
                'failures': 0,
                'processing_times': [],
                'employee_counts': [],
            },
            'backend_stability': {
                'recalculations': 0,
                'drift_detections': 0,
                'memory_leaks': [],
                'deadlocks': 0,
            },
            'frontend_sync': {
                'interactions': 0,
                'synchronization_errors': 0,
                'lag_events': [],
            },
            'export_validation': {
                'exports': 0,
                'csv_diffs': 0,
                'pdf_diffs': 0,
                'export_times': [],
            },
            'system_metrics': [],
            'errors': [],
        }
        
        # Frontend Playwright sessions are created per module run
    
    def _open_playwright_session(self) -> Tuple[Any, Any, Any, Any]:
        """Start a fresh Playwright session and ensure the context is authenticated."""
        if not PLAYWRIGHT_AVAILABLE:
            raise ImportError("Playwright package not installed. Install with: pip install playwright && playwright install chromium")
        
        playwright = None
        browser = None
        context = None
        page = None
        
        storage_state_path = (Path(__file__).resolve().parent.parent / 'auth' / 'state.json')

        try:
            print("   Starting fresh Playwright session...")
            playwright = sync_playwright().start()
            browser = playwright.chromium.launch(headless=True)
            if storage_state_path.exists():
                context = browser.new_context(storage_state=str(storage_state_path))
            else:
                context = browser.new_context()

            page = context.new_page()

            try:
                page.goto(f'{self.app_url}/dashboard', timeout=30000)
                page.wait_for_load_state('networkidle')
            except Exception:
                pass

            if '/dashboard' not in page.url:
                self._perform_playwright_login(page)
                if context:
                    try:
                        context.storage_state(path=str(storage_state_path))
                    except Exception:
                        pass
            return playwright, browser, context, page
        except Exception as exc:
            # Ensure everything is closed on failure
            print(f"   ❌ Failed to open Playwright session: {exc}")
            print(traceback.format_exc())
            self._close_playwright_session(playwright, browser, context, page)
            raise
    
    def _perform_playwright_login(self, page: Page) -> None:
        """Perform UI login within the provided Playwright page."""
        try:
            page.goto(f'{self.app_url}/login', timeout=30000)
            page.wait_for_load_state('networkidle')

            if re.search(r'/(dashboard|home)(/)?$', page.url):
                return

            if page.locator('input[name="username"]').count() > 0:
                page.fill('input[name="username"]', self.admin_username)

            if page.locator('input[name="password"]').count() == 0:
                raise RuntimeError('Password field not found on login page')

            page.fill('input[name="password"]', self.admin_password)
            page.click('button[type="submit"]')
            page.wait_for_url(re.compile(r'.*/(dashboard|home)(/)?$'), timeout=30000)
            page.wait_for_load_state('networkidle')
        except Exception as exc:
            raise RuntimeError(f'Playwright UI login failed: {exc}')

    def _close_playwright_session(self, playwright, browser, context, page) -> None:
        """Close Playwright resources safely."""
        for handle in [page, context, browser]:
            try:
                if handle:
                    handle.close()
            except Exception:
                pass
        if playwright:
            try:
                playwright.stop()
            except Exception:
                pass
    
    def test_excel_import_load(self, employee_counts: List[int] = [100, 250, 500]):
        """Module 1: Excel Import Load Testing"""
        print("\n" + "="*80)
        print("📊 MODULE 1: Excel Import Load Testing")
        print("="*80)
        
        for employee_count in employee_counts:
            try:
                # Generate Excel dataset
                excel_path = tempfile.mktemp(suffix='.xlsx')
                ExcelDatasetGenerator.generate_excel_dataset(excel_path, employee_count)
                
                # Measure import performance
                import_start = time.time()
                result = process_excel_file(excel_path)
                import_time = time.time() - import_start
                
                self.results['excel_import_tests']['runs'] += 1
                
                if result.get('success'):
                    self.results['excel_import_tests']['successes'] += 1
                    self.results['excel_import_tests']['processing_times'].append(import_time)
                    self.results['excel_import_tests']['employee_counts'].append(employee_count)
                    
                    print(f"✅ Imported {employee_count} employees in {import_time:.2f}s")
                    
                    # Verify imported totals
                    employees_processed = result.get('employees_processed', 0)
                    if employees_processed != employee_count:
                        error = {
                            'test': 'excel_import_count',
                            'expected': employee_count,
                            'actual': employees_processed,
                            'timestamp': datetime.now().isoformat()
                        }
                        self.results['errors'].append(error)
                        print(f"⚠️  Employee count mismatch: expected {employee_count}, got {employees_processed}")
                else:
                    self.results['excel_import_tests']['failures'] += 1
                    error = {
                        'test': 'excel_import',
                        'error': result.get('error', 'Unknown error'),
                        'timestamp': datetime.now().isoformat()
                    }
                    self.results['errors'].append(error)
                    print(f"❌ Import failed: {result.get('error')}")
                
                # Cleanup
                if os.path.exists(excel_path):
                    os.unlink(excel_path)
                    
            except Exception as e:
                self.results['excel_import_tests']['failures'] += 1
                error = {
                    'test': 'excel_import_exception',
                    'error': str(e),
                    'traceback': traceback.format_exc(),
                    'timestamp': datetime.now().isoformat()
                }
                self.results['errors'].append(error)
                print(f"❌ Excel import exception: {e}")
    
    def test_backend_stability(self):
        """Module 2: Backend Stability & Memory Integrity"""
        print("\n" + "="*80)
        print("🔧 MODULE 2: Backend Stability & Memory Integrity")
        print("="*80)
        
        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            # Get all employees
            cursor.execute('SELECT id, name FROM employees')
            employees = cursor.fetchall()
            
            if not employees:
                print("⚠️  No employees in database, skipping backend stability test")
                conn.close()
                return
            
            today = date.today()
            
            # Run continuous recalculations
            for employee in employees:
                employee_id = employee['id']
                employee_name = employee['name']
                
                # Get trips for employee
                cursor.execute('''
                    SELECT country, entry_date, exit_date, travel_days
                    FROM trips
                    WHERE employee_id = ?
                ''', (employee_id,))
                
                trips = [dict(row) for row in cursor.fetchall()]
                
                if not trips:
                    continue
                
                # Calculate days used multiple times to check for drift
                for iteration in range(10):
                    days_used = calculate_eu_days_from_trips(trips, today)
                    
                    # Check for drift
                    drift = self.drift_tracker.check_drift(employee_name, days_used)
                    if drift:
                        self.results['backend_stability']['drift_detections'] += 1
                        print(f"⚠️  Drift detected for {employee_name}: {drift['drift']}")
                    
                    if iteration == 0:
                        self.drift_tracker.record_baseline(employee_name, days_used)
                    
                    self.results['backend_stability']['recalculations'] += 1
            
            # Check for memory growth
            current_metric = self.monitor.record_metrics()
            self.results['system_metrics'].append(current_metric)
            
            # Check for SQLite lock issues
            try:
                cursor.execute('BEGIN IMMEDIATE')
                cursor.execute('SELECT 1')
                cursor.execute('COMMIT')
            except sqlite3.OperationalError as e:
                if 'locked' in str(e).lower():
                    self.results['backend_stability']['deadlocks'] += 1
                    print(f"⚠️  Database lock detected: {e}")
            
            conn.close()
            
        except Exception as e:
            error = {
                'test': 'backend_stability_exception',
                'error': str(e),
                'traceback': traceback.format_exc(),
                'timestamp': datetime.now().isoformat()
            }
            self.results['errors'].append(error)
            print(f"❌ Backend stability exception: {e}")
    
    
    def test_frontend_synchronization(self):
        """Module 3: Frontend Synchronization Stress"""
        print("\n" + "="*80)
        print("🖥️  MODULE 3: Frontend Synchronization Stress")
        print("="*80)

        session = None
        conn = None
        try:
            session = self._open_playwright_session()
            playwright, browser, context, page = session
        except Exception as exc:
            error = {
                'test': 'frontend_sync_initialization',
                'error': str(exc),
                'traceback': traceback.format_exc(),
                'timestamp': datetime.now().isoformat()
            }
            self.results['errors'].append(error)
            self.results['frontend_sync']['synchronization_errors'] += 1
            print(f"❌ Frontend sync initialization failed: {exc}")
            return

        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute('SELECT id, name FROM employees')
            employees = cursor.fetchall()

            page.goto(f'{self.app_url}/dashboard', timeout=30000)
            page.wait_for_load_state('networkidle')

            for interaction_num in range(10):
                interaction_start = time.time()

                try:
                    page.reload(timeout=30000)
                    page.wait_for_load_state('networkidle')
                except Exception as reload_error:
                    print(f"   ⚠️  Page reload error (interaction {interaction_num + 1}): {reload_error}")
                    # Attempt to restart session once
                    try:
                        self._close_playwright_session(*session)
                        session = None
                        session = self._open_playwright_session()
                        playwright, browser, context, page = session
                        page.goto(f'{self.app_url}/dashboard', timeout=30000)
                        page.wait_for_load_state('networkidle')
                    except Exception as restart_error:
                        print(f"   ❌ Unable to recover Playwright session: {restart_error}")
                        error = {
                            'test': 'frontend_sync_session_restart',
                            'error': str(restart_error),
                            'traceback': traceback.format_exc(),
                            'timestamp': datetime.now().isoformat()
                        }
                        self.results['errors'].append(error)
                        self.results['frontend_sync']['synchronization_errors'] += 1
                        break

                employee_rows = page.locator('tbody tr').all()

                for employee_row in employee_rows[:5]:
                    try:
                        employee_name_elem = employee_row.locator('td:first-child').first()
                        days_used_elem = employee_row.locator('td:nth-child(2), .days-used').first()

                        if employee_name_elem.count() > 0 and days_used_elem.count() > 0:
                            displayed_name = employee_name_elem.text_content()
                            displayed_days = days_used_elem.text_content()

                            for emp in employees:
                                if emp['name'] == displayed_name.strip():
                                    cursor.execute('''
                                        SELECT country, entry_date, exit_date
                                        FROM trips
                                        WHERE employee_id = ?
                                    ''', (emp['id'],))

                                    trips = [dict(row) for row in cursor.fetchall()]
                                    backend_days = calculate_eu_days_from_trips(trips, date.today())

                                    try:
                                        displayed_days_int = int(displayed_days.split('/')[0].strip())

                                        if abs(displayed_days_int - backend_days) > 0:
                                            error = {
                                                'test': 'frontend_backend_sync',
                                                'employee': displayed_name,
                                                'backend_days': backend_days,
                                                'frontend_days': displayed_days_int,
                                                'timestamp': datetime.now().isoformat()
                                            }
                                            self.results['frontend_sync']['synchronization_errors'] += 1
                                            self.results['errors'].append(error)
                                    except (ValueError, AttributeError):
                                        pass
                                    break
                    except Exception:
                        pass

                interaction_time = (time.time() - interaction_start) * 1000

                if interaction_time > SYNC_LAG_THRESHOLD_MS:
                    lag_event = {
                        'timestamp': datetime.now().isoformat(),
                        'lag_ms': interaction_time
                    }
                    self.results['frontend_sync']['lag_events'].append(lag_event)

                self.results['frontend_sync']['interactions'] += 1

        except Exception as e:
            error = {
                'test': 'frontend_sync_exception',
                'error': str(e),
                'traceback': traceback.format_exc(),
                'timestamp': datetime.now().isoformat()
            }
            self.results['errors'].append(error)
            print(f"❌ Frontend sync exception: {e}")
        finally:
            if conn:
                try:
                    conn.close()
                except Exception:
                    pass
            if session:
                self._close_playwright_session(*session)
    
    def test_export_validation(self):
        """Module 4: Export & Reporting Validation"""
        print("\n" + "="*80)
        print("📄 MODULE 4: Export & Reporting Validation")
        print("="*80)
        
        try:
            # Export CSV
            export_start = time.time()
            csv_data = export_trips_csv(self.db_path)
            export_time = time.time() - export_start
            
            self.results['export_validation']['exports'] += 1
            self.results['export_validation']['export_times'].append(export_time)
            
            # Export again and compare
            csv_data_2 = export_trips_csv(self.db_path)
            
            if csv_data != csv_data_2:
                self.results['export_validation']['csv_diffs'] += 1
                print("⚠️  CSV exports differ between calls")
                
                # Use diff library to show differences if available
                if DIFF_LIB_AVAILABLE:
                    dmp = diff_match_patch()
                    diffs = dmp.diff_main(csv_data, csv_data_2)
                    dmp.diff_cleanupSemantic(diffs)
                    diff_text = '\n'.join([f"{op}: {text[:50]}" for op, text in diffs[:10]])
                else:
                    # Simple comparison
                    diff_text = f"CSV length differs: {len(csv_data)} vs {len(csv_data_2)}"
                
                error = {
                    'test': 'csv_export_diff',
                    'diff': diff_text,
                    'timestamp': datetime.now().isoformat()
                }
                self.results['errors'].append(error)
            else:
                print(f"✅ CSV export consistent ({export_time:.3f}s)")
            
            # Export PDF
            pdf_start = time.time()
            pdf_data = export_all_employees_report_pdf(self.db_path)
            pdf_time = time.time() - pdf_start
            
            self.results['export_validation']['exports'] += 1
            self.results['export_validation']['export_times'].append(pdf_time)
            
            # Export PDF again and compare byte-level
            # Note: PDFs may contain timestamps, so we compare sizes and allow small differences
            time.sleep(0.1)  # Small delay to ensure any timestamp differences
            pdf_data_2 = export_all_employees_report_pdf(self.db_path)
            
            # PDFs often contain timestamps, so byte-level comparison may always differ
            # Instead, we check if sizes are similar (within 5% variance) and structure is intact
            size_diff = abs(len(pdf_data) - len(pdf_data_2))
            size_variance = size_diff / max(len(pdf_data), len(pdf_data_2), 1) * 100
            
            if pdf_data != pdf_data_2:
                # If variance is significant (>5%), log as potential issue
                if size_variance > 5.0:
                    self.results['export_validation']['pdf_diffs'] += 1
                    print(f"⚠️  PDF exports differ significantly (size variance: {size_variance:.2f}%)")
                    error = {
                        'test': 'pdf_export_diff',
                        'size1': len(pdf_data),
                        'size2': len(pdf_data_2),
                        'size_diff': size_diff,
                        'variance_percent': size_variance,
                        'timestamp': datetime.now().isoformat()
                    }
                    self.results['errors'].append(error)
                else:
                    # Small differences likely due to timestamps - this is acceptable
                    print(f"✅ PDF export consistent ({pdf_time:.3f}s, minor timestamp differences expected)")
            else:
                print(f"✅ PDF export identical ({pdf_time:.3f}s)")
            
        except Exception as e:
            error = {
                'test': 'export_validation_exception',
                'error': str(e),
                'traceback': traceback.format_exc(),
                'timestamp': datetime.now().isoformat()
            }
            self.results['errors'].append(error)
            print(f"❌ Export validation exception: {e}")
    
    def run_endurance_test(self):
        """Run the full 180-minute endurance test"""
        self.start_time = time.time()
        self.end_time = self.start_time + TEST_DURATION_SECONDS
        
        self.results['start_time'] = datetime.now().isoformat()
        
        print("\n" + "="*80)
        print("🌙 PHASE 3.10: ELITE QA ENDURANCE & STRESS TESTING")
        print("="*80)
        print(f"Test Duration: {TEST_DURATION_MINUTES} minutes ({TEST_DURATION_SECONDS} seconds)")
        print(f"Start Time: {self.results['start_time']}")
        print(f"Expected End Time: {datetime.fromtimestamp(self.end_time).isoformat()}")
        print("="*80)
        
        cycle_count = 0
        
        try:
            while time.time() < self.end_time:
                cycle_count += 1
                elapsed = time.time() - self.start_time
                remaining = self.end_time - time.time()
                
                print(f"\n{'='*80}")
                print(f"Cycle {cycle_count} | Elapsed: {elapsed/3600:.2f}h | Remaining: {remaining/3600:.2f}h")
                print(f"{'='*80}")
                
                # Run all test modules
                self.test_excel_import_load([100, 250, 500])
                self.test_backend_stability()
                self.test_frontend_synchronization()
                self.test_export_validation()
                
                # Record system metrics every minute
                if cycle_count % 10 == 0:  # Every 10 cycles (adjust based on cycle time)
                    self.monitor.record_metrics()
                
                # Brief pause between cycles
                time.sleep(5)
        
        finally:
            self.end_time = time.time()
            self.results['end_time'] = datetime.now().isoformat()
            self.results['duration_seconds'] = self.end_time - self.start_time
            
            # Final metrics
            self.monitor.record_metrics()
            self.results['system_metrics'] = self.monitor.get_metrics()
            
            # Generate reports
            self.generate_reports()
            
            print("\n" + "="*80)
            print("✅ ENDURANCE TEST COMPLETE")
            print("="*80)
            print(f"Duration: {self.results['duration_seconds']/3600:.2f} hours")
            print(f"Total Errors: {len(self.results['errors'])}")
            print(f"Drift Detections: {self.results['backend_stability']['drift_detections']}")
            print(f"Sync Errors: {self.results['frontend_sync']['synchronization_errors']}")
            print(f"Export Diffs: {self.results['export_validation']['csv_diffs'] + self.results['export_validation']['pdf_diffs']}")
            print("="*80)
    
    def generate_reports(self):
        """Generate all required output reports"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # 1. HTML Report
        html_path = REPORTS_DIR / f'qa_endurance_report_{timestamp}.html'
        self._generate_html_report(html_path)
        
        # 2. JSON Metrics
        json_path = REPORTS_DIR / f'qa_endurance_metrics_{timestamp}.json'
        self._generate_json_metrics(json_path)
        
        # 3. Memory Graph
        graph_path = REPORTS_DIR / f'qa_memory_graph_{timestamp}.png'
        self._generate_memory_graph(graph_path)
        
        # 4. Drift Summary
        drift_path = REPORTS_DIR / f'qa_drift_summary_{timestamp}.txt'
        self._generate_drift_summary(drift_path)
        
        # 5. Export Diff Log
        export_diff_path = REPORTS_DIR / f'qa_export_diff_{timestamp}.log'
        self._generate_export_diff_log(export_diff_path)
        
        print(f"\n📊 Reports generated:")
        print(f"   - {html_path}")
        print(f"   - {json_path}")
        print(f"   - {graph_path}")
        print(f"   - {drift_path}")
        print(f"   - {export_diff_path}")
    
    def _generate_html_report(self, output_path: Path):
        """Generate HTML summary dashboard"""
        html = f"""<!DOCTYPE html>
<html>
<head>
    <title>Phase 3.10 Endurance Test Report - ComplyEur</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 20px; border-radius: 8px; }}
        h1 {{ color: #2c5282; }}
        .metric {{ background: #f7fafc; padding: 15px; margin: 10px 0; border-radius: 4px; }}
        .success {{ color: green; }}
        .warning {{ color: orange; }}
        .error {{ color: red; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        th, td {{ padding: 10px; text-align: left; border-bottom: 1px solid #ddd; }}
        th {{ background: #2c5282; color: white; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🌙 Phase 3.10: Elite QA Endurance & Stress Test Report</h1>
        <p><strong>ComplyEur</strong> - EU 90/180 Employee Travel Tracker</p>
        
        <div class="metric">
            <h2>Test Summary</h2>
            <p><strong>Start Time:</strong> {self.results['start_time']}</p>
            <p><strong>End Time:</strong> {self.results['end_time']}</p>
            <p><strong>Duration:</strong> {self.results['duration_seconds']/3600:.2f} hours</p>
        </div>
        
        <div class="metric">
            <h2>Excel Import Load Testing</h2>
            <p><strong>Total Runs:</strong> {self.results['excel_import_tests']['runs']}</p>
            <p><strong>Successes:</strong> <span class="success">{self.results['excel_import_tests']['successes']}</span></p>
            <p><strong>Failures:</strong> <span class="error">{self.results['excel_import_tests']['failures']}</span></p>
            {f"<p><strong>Avg Processing Time:</strong> {sum(self.results['excel_import_tests']['processing_times'])/len(self.results['excel_import_tests']['processing_times']):.2f}s</p>" if self.results['excel_import_tests']['processing_times'] else ""}
        </div>
        
        <div class="metric">
            <h2>Backend Stability</h2>
            <p><strong>Recalculations:</strong> {self.results['backend_stability']['recalculations']}</p>
            <p><strong>Drift Detections:</strong> <span class="{'error' if self.results['backend_stability']['drift_detections'] > 0 else 'success'}">{self.results['backend_stability']['drift_detections']}</span></p>
            <p><strong>Deadlocks:</strong> <span class="{'error' if self.results['backend_stability']['deadlocks'] > 0 else 'success'}">{self.results['backend_stability']['deadlocks']}</span></p>
        </div>
        
        <div class="metric">
            <h2>Frontend Synchronization</h2>
            <p><strong>Interactions:</strong> {self.results['frontend_sync']['interactions']}</p>
            <p><strong>Synchronization Errors:</strong> <span class="{'error' if self.results['frontend_sync']['synchronization_errors'] > 0 else 'success'}">{self.results['frontend_sync']['synchronization_errors']}</span></p>
            <p><strong>Lag Events (>50ms):</strong> {len(self.results['frontend_sync']['lag_events'])}</p>
        </div>
        
        <div class="metric">
            <h2>Export Validation</h2>
            <p><strong>Total Exports:</strong> {self.results['export_validation']['exports']}</p>
            <p><strong>CSV Diffs:</strong> <span class="{'error' if self.results['export_validation']['csv_diffs'] > 0 else 'success'}">{self.results['export_validation']['csv_diffs']}</span></p>
            <p><strong>PDF Diffs:</strong> <span class="{'error' if self.results['export_validation']['pdf_diffs'] > 0 else 'success'}">{self.results['export_validation']['pdf_diffs']}</span></p>
        </div>
        
        <div class="metric">
            <h2>System Metrics Summary</h2>
            {self._generate_system_metrics_html()}
        </div>
        
        <div class="metric">
            <h2>Errors & Issues</h2>
            <p><strong>Total Errors:</strong> <span class="{'error' if len(self.results['errors']) > 0 else 'success'}">{len(self.results['errors'])}</span></p>
            {self._generate_errors_html()}
        </div>
    </div>
</body>
</html>"""
        
        with open(output_path, 'w') as f:
            f.write(html)
    
    def _generate_system_metrics_html(self) -> str:
        """Generate HTML for system metrics"""
        if not self.results['system_metrics']:
            return "<p>No metrics recorded</p>"
        
        summary = self.monitor.get_summary()
        if not summary:
            return "<p>No metrics summary available</p>"
        
        return f"""
        <table>
            <tr><th>Metric</th><th>Value</th></tr>
            <tr><td>CPU Average</td><td>{summary.get('cpu_avg', 0):.2f}%</td></tr>
            <tr><td>CPU Max</td><td>{summary.get('cpu_max', 0):.2f}%</td></tr>
            <tr><td>Memory Average</td><td>{summary.get('memory_avg_mb', 0):.2f} MB</td></tr>
            <tr><td>Memory Max</td><td>{summary.get('memory_max_mb', 0):.2f} MB</td></tr>
            <tr><td>Memory Growth</td><td>{summary.get('memory_growth_mb', 0):.2f} MB</td></tr>
            <tr><td>Total I/O Read</td><td>{summary.get('total_io_read_mb', 0):.2f} MB</td></tr>
            <tr><td>Total I/O Write</td><td>{summary.get('total_io_write_mb', 0):.2f} MB</td></tr>
        </table>
        """
    
    def _generate_errors_html(self) -> str:
        """Generate HTML for errors"""
        if not self.results['errors']:
            return "<p class='success'>✅ No errors detected</p>"
        
        error_rows = ""
        for error in self.results['errors'][:20]:  # Show first 20
            error_rows += f"<tr><td>{error.get('test', 'Unknown')}</td><td>{error.get('error', 'N/A')}</td><td>{error.get('timestamp', 'N/A')}</td></tr>"
        
        return f"""
        <table>
            <tr><th>Test</th><th>Error</th><th>Timestamp</th></tr>
            {error_rows}
        </table>
        <p><em>Showing first 20 errors. Total: {len(self.results['errors'])}</em></p>
        """
    
    def _generate_json_metrics(self, output_path: Path):
        """Generate JSON metrics file"""
        metrics_data = {
            'test_info': {
                'start_time': self.results['start_time'],
                'end_time': self.results['end_time'],
                'duration_seconds': self.results['duration_seconds'],
                'duration_hours': self.results['duration_seconds'] / 3600,
            },
            'results': self.results,
            'system_summary': self.monitor.get_summary(),
            'drift_log': self.drift_tracker.get_drift_log(),
        }
        
        with open(output_path, 'w') as f:
            json.dump(metrics_data, f, indent=2)
    
    def _generate_memory_graph(self, output_path: Path):
        """Generate memory stability plot"""
        try:
            import matplotlib
            matplotlib.use('Agg')  # Non-interactive backend
            import matplotlib.pyplot as plt
            
            if not self.results['system_metrics']:
                # Create empty graph
                plt.figure(figsize=(12, 6))
                plt.text(0.5, 0.5, 'No metrics recorded', ha='center', va='center')
                plt.savefig(output_path)
                plt.close()
                return
            
            timestamps = [m['elapsed_seconds'] / 3600 for m in self.results['system_metrics']]
            memory_values = [m['memory_mb'] for m in self.results['system_metrics']]
            
            plt.figure(figsize=(12, 6))
            plt.plot(timestamps, memory_values, 'b-', linewidth=2)
            plt.xlabel('Time (hours)')
            plt.ylabel('Memory Usage (MB)')
            plt.title('Memory Stability Over Test Duration')
            plt.grid(True, alpha=0.3)
            plt.savefig(output_path, dpi=150, bbox_inches='tight')
            plt.close()
            
        except ImportError:
            # Matplotlib not available, create text file instead
            with open(output_path.with_suffix('.txt'), 'w') as f:
                f.write("Memory Graph (matplotlib not available)\n")
                f.write("="*50 + "\n")
                for metric in self.results['system_metrics']:
                    f.write(f"Time: {metric['elapsed_seconds']:.1f}s | Memory: {metric['memory_mb']:.2f} MB\n")
        except Exception as e:
            print(f"⚠️  Could not generate memory graph: {e}")
    
    def _generate_drift_summary(self, output_path: Path):
        """Generate drift summary text file"""
        drift_log = self.drift_tracker.get_drift_log()
        
        with open(output_path, 'w') as f:
            f.write("="*80 + "\n")
            f.write("FLOATING-POINT DRIFT SUMMARY\n")
            f.write("="*80 + "\n\n")
            f.write(f"Tolerance: {DRIFT_TOLERANCE}\n")
            f.write(f"Total Drift Detections: {len(drift_log)}\n\n")
            
            if drift_log:
                f.write("Drift Events:\n")
                f.write("-"*80 + "\n")
                for drift in drift_log:
                    f.write(f"Timestamp: {drift['timestamp']}\n")
                    f.write(f"Employee: {drift['employee']}\n")
                    f.write(f"Baseline: {drift['baseline']}\n")
                    f.write(f"Current: {drift['current']}\n")
                    f.write(f"Drift: {drift['drift']}\n")
                    f.write("-"*80 + "\n")
            else:
                f.write("✅ No drift detected - all calculations remained stable.\n")
    
    def _generate_export_diff_log(self, output_path: Path):
        """Generate export diff log"""
        export_errors = [e for e in self.results['errors'] if 'export' in e.get('test', '').lower()]
        
        with open(output_path, 'w') as f:
            f.write("="*80 + "\n")
            f.write("EXPORT VERIFICATION RESULTS\n")
            f.write("="*80 + "\n\n")
            f.write(f"Total Exports: {self.results['export_validation']['exports']}\n")
            f.write(f"CSV Diffs: {self.results['export_validation']['csv_diffs']}\n")
            f.write(f"PDF Diffs: {self.results['export_validation']['pdf_diffs']}\n\n")
            
            if export_errors:
                f.write("Export Differences:\n")
                f.write("-"*80 + "\n")
                for error in export_errors:
                    f.write(f"Test: {error.get('test')}\n")
                    f.write(f"Timestamp: {error.get('timestamp')}\n")
                    if 'diff' in error:
                        f.write(f"Diff: {error['diff']}\n")
                    if 'size1' in error:
                        f.write(f"Size 1: {error['size1']} bytes\n")
                        f.write(f"Size 2: {error['size2']} bytes\n")
                    f.write("-"*80 + "\n")
            else:
                f.write("✅ All exports verified - no discrepancies detected.\n")


def main():
    """Main entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Phase 3.10 Endurance & Stress Testing')
    parser.add_argument('--db-path', type=str, default='data/eu_tracker.db',
                       help='Path to SQLite database')
    parser.add_argument('--app-url', type=str, default='http://127.0.0.1:5001',
                       help='Flask application URL')
    parser.add_argument('--admin-password', type=str, default='admin123',
                       help='Admin password for login')
    parser.add_argument('--admin-username', type=str, default='admin',
                       help='Admin username for login')
    parser.add_argument('--duration', type=int, default=None,
                       help='Test duration in minutes (overrides TEST_DURATION_MINUTES env var)')
    
    args = parser.parse_args()
    
    # Override duration if provided
    global TEST_DURATION_MINUTES, TEST_DURATION_SECONDS
    if args.duration:
        TEST_DURATION_MINUTES = args.duration
        TEST_DURATION_SECONDS = TEST_DURATION_MINUTES * 60
    
    # Ensure database exists
    if not os.path.exists(args.db_path):
        print(f"⚠️  Database not found at {args.db_path}")
        print("   Creating test database...")
        
        # Create database directory if needed
        os.makedirs(os.path.dirname(args.db_path), exist_ok=True)
        
        # Initialize database
        conn = sqlite3.connect(args.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS employees (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL
            )
        ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS trips (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER,
                country TEXT,
                entry_date DATE,
                exit_date DATE,
                travel_days INTEGER DEFAULT 0,
                FOREIGN KEY(employee_id) REFERENCES employees(id)
            )
        ''')
        conn.commit()
        conn.close()
    
    # Run test suite
    suite = EnduranceStressTestSuite(
        db_path=args.db_path,
        app_url=args.app_url,
        admin_password=args.admin_password,
        admin_username=args.admin_username
    )
    
    suite.run_endurance_test()


if __name__ == '__main__':
    main()

