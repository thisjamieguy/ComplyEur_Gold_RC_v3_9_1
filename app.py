from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, make_response, send_file
import hashlib
import sqlite3
from datetime import datetime, timedelta
from functools import wraps
import os
import re
import openpyxl
from werkzeug.utils import secure_filename
from config import load_config, get_session_lifetime, save_config
from modules.services.hashing import Hasher
from modules.services.audit import write_audit
from modules.services.retention import purge_expired_trips, get_expired_trips, anonymize_employee
from modules.services.dsar import create_dsar_export, delete_employee_data, rectify_employee_name, get_employee_data
from modules.services.exports import export_trips_csv, export_employee_report_pdf, export_all_employees_report_pdf
from modules.services.rolling90 import presence_days, days_used_in_window, earliest_safe_entry, calculate_days_remaining, get_risk_level, days_until_compliant
from modules.services.trip_validator import validate_trip, validate_date_range
import io
import csv
import zipfile
import secrets
import json
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Application Version
APP_VERSION = "1.1 (post-security review)"

app = Flask(__name__)

# ================================================================================
# TEMPORARY TEST ROUTE - REMOVE WHEN TESTING COMPLETE
# ================================================================================
# Import and register test overview blueprint
# TO REMOVE: Delete these 3 lines and delete test_overview.py file
from test_overview import test_bp
app.register_blueprint(test_bp)
# ================================================================================

# Load EU Entry Requirements data at startup
EU_ENTRY_DATA = []
try:
    data_file_path = os.path.join(os.path.dirname(__file__), 'data', 'eu_entry_requirements.json')
    with open(data_file_path, 'r', encoding='utf-8') as f:
        EU_ENTRY_DATA = json.load(f)
    print(f"✓ Loaded {len(EU_ENTRY_DATA)} EU entry requirements")
except Exception as e:
    print(f"Warning: Could not load EU entry requirements data: {e}")

# Secure secret key from environment variable
SECRET_KEY = os.getenv('SECRET_KEY')
if not SECRET_KEY or SECRET_KEY == 'your-secret-key-here-change-this-to-a-random-64-character-hex-string':
    # Generate a secure random key if not set
    SECRET_KEY = secrets.token_hex(32)
    print("\n" + "="*70)
    print("WARNING: No SECRET_KEY found in .env file!")
    print("Using a temporary key. Add this to your .env file:")
    print(f"SECRET_KEY={SECRET_KEY}")
    print("="*70 + "\n")

app.secret_key = SECRET_KEY
CONFIG = load_config()

# Configure static files
app.static_folder = 'static'
DATABASE = os.path.join(os.path.dirname(__file__), os.getenv('DATABASE_PATH', 'eu_tracker.db'))

# Make version available to all templates
@app.context_processor
def inject_version():
    return dict(app_version=APP_VERSION)

# Session cookie hardening (essential-only cookie)
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',  # Changed to 'Lax' for better compatibility
    SESSION_COOKIE_SECURE=os.getenv('SESSION_COOKIE_SECURE', 'False').lower() == 'true'
)
app.permanent_session_lifetime = get_session_lifetime(CONFIG)

# Additional security headers
@app.after_request
def set_security_headers(response):
    """Add security headers to all responses"""
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    # Only add HSTS if using HTTPS
    if request.is_secure:
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    return response

# Upload configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Ensure upload folder exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Schengen country code mapping
COUNTRY_CODE_MAPPING = {
    'AT': 'Austria',
    'BE': 'Belgium',
    'BG': 'Bulgaria',
    'HR': 'Croatia',
    'CY': 'Cyprus',
    'CZ': 'Czech Republic',
    'DK': 'Denmark',
    'EE': 'Estonia',
    'FI': 'Finland',
    'FR': 'France',
    'DE': 'Germany',
    'GR': 'Greece',
    'HU': 'Hungary',
    'IS': 'Iceland',
    'IE': 'Ireland',
    'IT': 'Italy',
    'LV': 'Latvia',
    'LI': 'Liechtenstein',
    'LT': 'Lithuania',
    'LU': 'Luxembourg',
    'MT': 'Malta',
    'NL': 'Netherlands',
    'NO': 'Norway',
    'PL': 'Poland',
    'PT': 'Portugal',
    'RO': 'Romania',
    'SK': 'Slovakia',
    'SI': 'Slovenia',
    'ES': 'Spain',
    'SE': 'Sweden',
    'CH': 'Switzerland'
}

# Reverse lookup
COUNTRY_NAME_TO_CODE = {v: k for k, v in COUNTRY_CODE_MAPPING.items()}

hasher = Hasher(os.getenv('PASSWORD_HASH_SCHEME', 'argon2'))
LOGIN_ATTEMPTS = {}
MAX_ATTEMPTS = int(os.getenv('MAX_LOGIN_ATTEMPTS', 5))
ATTEMPT_WINDOW_SECONDS = int(os.getenv('ATTEMPT_WINDOW_SECONDS', 600))

# ================================================================================
# INPUT VALIDATION & SANITIZATION
# ================================================================================

def validate_employee_name(name):
    """Validate and sanitize employee name input"""
    if not name or not isinstance(name, str):
        return None, "Employee name is required"
    
    # Remove dangerous characters
    name = re.sub(r'[<>"\'&;]', '', name.strip())
    
    # Check length
    if len(name) < 2:
        return None, "Employee name must be at least 2 characters"
    if len(name) > 100:
        return None, "Employee name must be less than 100 characters"
    
    # Check for valid characters (letters, spaces, hyphens, apostrophes)
    if not re.match(r'^[a-zA-Z\s\-\']+$', name):
        return None, "Employee name contains invalid characters"
    
    return name, None

def validate_date(date_str, field_name="Date"):
    """Validate date format and return date object"""
    if not date_str:
        return None, f"{field_name} is required"
    
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # Check reasonable date range (not too far in past or future)
        today = datetime.now().date()
        min_date = today - timedelta(days=3650)  # 10 years ago
        max_date = today + timedelta(days=730)   # 2 years in future
        
        if date_obj < min_date:
            return None, f"{field_name} is too far in the past"
        if date_obj > max_date:
            return None, f"{field_name} is too far in the future"
        
        return date_obj, None
    except ValueError:
        return None, f"Invalid {field_name.lower()} format. Use YYYY-MM-DD"

def validate_date_range(entry_date, exit_date):
    """Validate that exit date is after entry date"""
    if exit_date <= entry_date:
        return False, "Exit date must be after entry date"
    
    # Check trip duration is reasonable (max 2 years)
    if (exit_date - entry_date).days > 730:
        return False, "Trip duration cannot exceed 2 years"
    
    return True, None

def validate_country_code(country_code):
    """Validate country code"""
    if not country_code:
        return None, "Country code is required"
    
    # Sanitize and validate
    country_code = re.sub(r'[^A-Z]', '', str(country_code).upper().strip())
    
    if country_code not in COUNTRY_CODE_MAPPING:
        return None, f"Invalid country code: {country_code}"
    
    return country_code, None

def sanitize_string(value, max_length=255):
    """Generic string sanitization"""
    if not value:
        return ""
    
    # Remove dangerous characters
    value = re.sub(r'[<>"\'&;]', '', str(value).strip())
    return value[:max_length]

@app.template_filter('uk_date')
def uk_date(value):
    """Format YYYY-MM-DD or datetime to DD-MM-YYYY for UK display."""
    try:
        if not value:
            return value
        if isinstance(value, datetime):
            return value.strftime('%d-%m-%Y')
        if isinstance(value, str):
            try:
                dt = datetime.strptime(value, '%Y-%m-%d')
                return dt.strftime('%d-%m-%Y')
            except ValueError:
                return value
        return value
    except Exception:
        return value

@app.template_filter('country_name')
def country_name(value):
    try:
        if not value:
            return value
        code = str(value).upper()
        return COUNTRY_CODE_MAPPING.get(code, code)
    except Exception:
        return value

@app.template_filter('is_ireland')
def is_ireland(country):
    """Check if country is Ireland (Non-Schengen)"""
    if not country:
        return False
    country_upper = str(country).upper().strip()
    return country_upper in ['IE', 'IRELAND']

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def check_password(stored_hash, password):
    return stored_hash == hash_password(password)

def verify_and_upgrade_password(stored_hash: str, candidate: str):
    """Verify password against stored hash, supporting legacy SHA256.
    Returns (is_valid: bool, upgraded_hash_or_None)."""
    try:
        if hasher.verify(stored_hash, candidate):
            return True, None
    except Exception:
        pass
    # legacy sha256
    try:
        if stored_hash == hashlib.sha256(candidate.encode()).hexdigest():
            try:
                return True, hasher.hash(candidate)
            except Exception:
                # If upgrade hashing is unavailable, still allow login without upgrading
                return True, None
    except Exception:
        pass
    return False, None

def init_db():
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    
    # Enable foreign keys
    c.execute('PRAGMA foreign_keys = ON')
    
    # Create tables
    c.execute('CREATE TABLE IF NOT EXISTS admin (id INTEGER PRIMARY KEY, password_hash TEXT NOT NULL)')
    c.execute('CREATE TABLE IF NOT EXISTS employees (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL UNIQUE)')
    c.execute('''CREATE TABLE IF NOT EXISTS trips (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        employee_id INTEGER NOT NULL,
        country TEXT NOT NULL,
        entry_date TEXT NOT NULL,
        exit_date TEXT NOT NULL,
        travel_days INTEGER DEFAULT 0,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE
    )''')
    
    # Lightweight migrations for legacy schemas
    try:
        c.execute("PRAGMA table_info(trips)")
        cols = {row[1] for row in c.fetchall()}
        if 'country' not in cols:
            c.execute('ALTER TABLE trips ADD COLUMN country TEXT')
            if 'country_code' in cols:
                c.execute('UPDATE trips SET country = country_code WHERE country IS NULL')
        c.execute("PRAGMA table_info(trips)")
        cols = {row[1] for row in c.fetchall()}
        if 'travel_days' not in cols:
            c.execute('ALTER TABLE trips ADD COLUMN travel_days INTEGER DEFAULT 0')
        c.execute("PRAGMA table_info(trips)")
        cols = {row[1] for row in c.fetchall()}
        if 'created_at' not in cols:
            c.execute("ALTER TABLE trips ADD COLUMN created_at TEXT DEFAULT (datetime('now'))")
        c.execute("PRAGMA table_info(trips)")
        cols = {row[1] for row in c.fetchall()}
        if 'validation_note' not in cols:
            c.execute('ALTER TABLE trips ADD COLUMN validation_note TEXT')
    except Exception:
        pass
    
    # Create indexes for performance and GDPR compliance
    try:
        # Index on employee_id for faster joins and lookups
        c.execute('CREATE INDEX IF NOT EXISTS idx_trips_employee_id ON trips(employee_id)')
        # Index on exit_date for retention policy queries
        c.execute('CREATE INDEX IF NOT EXISTS idx_trips_exit_date ON trips(exit_date)')
        # Composite index on employee_id + entry_date for sorting/filtering
        c.execute('CREATE INDEX IF NOT EXISTS idx_trips_employee_entry ON trips(employee_id, entry_date)')
        # Index on entry_date for date range queries
        c.execute('CREATE INDEX IF NOT EXISTS idx_trips_entry_date ON trips(entry_date)')
    except Exception as e:
        print(f"Warning: Could not create indexes: {e}")
    
    # Initialize admin account if not exists
    c.execute('SELECT COUNT(*) FROM admin')
    if c.fetchone()[0] == 0:
        # Use password from environment variable or default
        default_password = os.getenv('ADMIN_PASSWORD', 'admin123')
        c.execute('INSERT INTO admin (password_hash) VALUES (?)', (hasher.hash(default_password),))
        print(f"\nAdmin account created with password from .env file")
        print("Please change the admin password immediately after first login!\n")
    
    conn.commit()
    conn.close()

def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.before_request
def enforce_session_idle_timeout():
    try:
        if session.get('logged_in'):
            now = datetime.utcnow()
            last = session.get('last_activity')
            if last:
                last_dt = datetime.strptime(last, '%Y-%m-%dT%H:%M:%SZ')
                if now - last_dt > get_session_lifetime(CONFIG):
                    session.clear()
                    flash('Session expired. Please log in again.')
                    return redirect(url_for('login'))
            session['last_activity'] = now.strftime('%Y-%m-%dT%H:%M:%SZ')
    except Exception:
        pass

def calculate_eu_days(employee_id):
    """Calculate Schengen days for employee using rolling90 logic (Ireland excluded)."""
    conn = get_db()
    c = conn.cursor()
    today = datetime.now().date()
    
    # Get all trips for this employee (including country info)
    c.execute('SELECT entry_date, exit_date, country FROM trips WHERE employee_id = ?', (employee_id,))
    trips = [{'entry_date': t['entry_date'], 'exit_date': t['exit_date'], 'country': t['country']} for t in c.fetchall()]
    
    # Use the new rolling90 logic that filters out Ireland
    presence = presence_days(trips)
    days_used = days_used_in_window(presence, today)
    
    conn.close()
    return days_used

def calculate_eu_days_at_date(trips_for_employee, reference_date):
    """Calculate Schengen days used within the 180-day window ending at reference_date.
    Ireland trips are excluded from calculations.
    trips_for_employee: list of dicts with 'entry_date', 'exit_date', and 'country'
    reference_date: datetime.date
    """
    # Use the new rolling90 logic that filters out Ireland
    presence = presence_days(trips_for_employee)
    return days_used_in_window(presence, reference_date)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def parse_schedule_excel_file(filepath):
    """Parse work schedule Excel file and extract EU trips based on country codes"""
    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active

        trips = []
        errors = []

        # Get employee names from first column (skip rows 4-9, and 68+)
        employees = {}
        for row_idx in range(10, min(68, sheet.max_row + 1)):  # Rows 10-67
            name_cell = sheet.cell(row=row_idx, column=1).value
            if name_cell and str(name_cell).strip() not in ['None', 'Unallocated', '']:
                employees[row_idx] = str(name_cell).strip()

        print(f"Found {len(employees)} employees in rows 10-67")

        # Get date headers from row 3 (starting from column 2)
        date_columns = {}
        for col_idx in range(2, sheet.max_column + 1):
            date_cell = sheet.cell(row=3, column=col_idx).value
            if date_cell and isinstance(date_cell, datetime):
                date_columns[col_idx] = date_cell.date()

        print(f"Found {len(date_columns)} date columns")

        # All possible EU/Schengen country codes
        eu_country_codes = list(COUNTRY_CODE_MAPPING.keys())

        # Process each employee's schedule
        for emp_row, emp_name in employees.items():
            current_country = None
            trip_start = None

            for col_idx in sorted(date_columns.keys()):
                date = date_columns[col_idx]
                cell_value = sheet.cell(row=emp_row, column=col_idx).value

                # Clean cell value
                assignment = None
                if cell_value:
                    assignment = str(cell_value).strip()
                    if assignment in ['None', '', ' ']:
                        assignment = None

                # Look for country codes in the assignment
                found_country = None
                if assignment:
                    assignment_upper = assignment.upper()
                    for country_code in eu_country_codes:
                        if country_code in assignment_upper:
                            found_country = country_code
                            break

                # If country changes, process the previous trip
                if found_country != current_country:
                    if current_country and trip_start:
                        # End current trip (yesterday)
                        trip_end = date_columns.get(col_idx - 1, date)

                        trips.append({
                            'name': emp_name,
                            'country': COUNTRY_CODE_MAPPING[current_country],
                            'entry_date': trip_start.strftime('%Y-%m-%d'),
                            'exit_date': trip_end.strftime('%Y-%m-%d'),
                            'row': emp_row
                        })

                    # Start new trip if we found a country
                    current_country = found_country
                    trip_start = date if found_country else None

            # Handle trip that extends to end of schedule
            if current_country and trip_start:
                last_date = max(date_columns.values())
                trips.append({
                    'name': emp_name,
                    'country': COUNTRY_CODE_MAPPING[current_country],
                    'entry_date': trip_start.strftime('%Y-%m-%d'),
                    'exit_date': last_date.strftime('%Y-%m-%d'),
                    'row': emp_row
                })

        return trips, errors

    except Exception as e:
        return [], [f"Error reading schedule Excel file: {str(e)}"]

def parse_excel_file(filepath):
    """Parse Excel file and extract trip data - detects format automatically"""
    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active

        # Detect if this is a schedule format vs. regular trip data format
        # Schedule format has: dates in row 3, employee names starting around row 10
        is_schedule_format = False

        # Check if row 3 has dates
        row3_has_dates = False
        for col_idx in range(2, min(10, sheet.max_column + 1)):
            cell_value = sheet.cell(row=3, column=col_idx).value
            if isinstance(cell_value, datetime):
                row3_has_dates = True
                break

        # Check if we have employee names around row 10-20
        has_employee_names = False
        for row_idx in range(10, min(21, sheet.max_row + 1)):
            name_cell = sheet.cell(row=row_idx, column=1).value
            if name_cell and isinstance(name_cell, str) and len(name_cell.strip()) > 3:
                has_employee_names = True
                break

        if row3_has_dates and has_employee_names:
            print("Detected schedule format - using schedule parser")
            return parse_schedule_excel_file(filepath)
        else:
            print("Detected standard format - using standard parser")
            # Continue with standard parsing logic

        trips = []
        errors = []

        # Try to detect headers
        headers = {}
        for col_idx, cell in enumerate(sheet[1], 1):
            if cell.value:
                header = str(cell.value).lower().strip()
                if 'name' in header or 'employee' in header:
                    headers['name'] = col_idx
                elif 'country' in header or 'destination' in header:
                    headers['country'] = col_idx
                elif 'entry' in header or 'start' in header or 'arrival' in header:
                    headers['entry_date'] = col_idx
                elif 'exit' in header or 'end' in header or 'departure' in header:
                    headers['exit_date'] = col_idx

        # If headers not found, assume standard order: Name, Country, Entry, Exit
        if not headers:
            headers = {'name': 1, 'country': 2, 'entry_date': 3, 'exit_date': 4}

        # Process data rows (starting from row 2)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):  # Skip empty rows
                continue

            try:
                # Extract data based on detected headers
                name = str(row[headers.get('name', 1) - 1] or '').strip() if len(row) >= headers.get('name', 1) else ''
                country_code = str(row[headers.get('country', 2) - 1] or '').strip().upper() if len(row) >= headers.get('country', 2) else ''
                entry_date = row[headers.get('entry_date', 3) - 1] if len(row) >= headers.get('entry_date', 3) else None
                exit_date = row[headers.get('exit_date', 4) - 1] if len(row) >= headers.get('exit_date', 4) else None

                # Validate required fields
                if not name:
                    errors.append(f"Row {row_idx}: Missing employee name")
                    continue

                if not country_code:
                    errors.append(f"Row {row_idx}: Missing country code")
                    continue

                # Convert country code to full name
                country_name = COUNTRY_CODE_MAPPING.get(country_code)
                if not country_name:
                    errors.append(f"Row {row_idx}: Unknown country code '{country_code}'")
                    continue

                # Parse dates
                if isinstance(entry_date, datetime):
                    entry_date_str = entry_date.strftime('%Y-%m-%d')
                elif isinstance(entry_date, str):
                    try:
                        parsed_date = datetime.strptime(entry_date.strip(), '%Y-%m-%d')
                        entry_date_str = parsed_date.strftime('%Y-%m-%d')
                    except ValueError:
                        try:
                            parsed_date = datetime.strptime(entry_date.strip(), '%d/%m/%Y')
                            entry_date_str = parsed_date.strftime('%Y-%m-%d')
                        except ValueError:
                            errors.append(f"Row {row_idx}: Invalid entry date format '{entry_date}'")
                            continue
                else:
                    errors.append(f"Row {row_idx}: Missing entry date")
                    continue

                if isinstance(exit_date, datetime):
                    exit_date_str = exit_date.strftime('%Y-%m-%d')
                elif isinstance(exit_date, str):
                    try:
                        parsed_date = datetime.strptime(exit_date.strip(), '%Y-%m-%d')
                        exit_date_str = parsed_date.strftime('%Y-%m-%d')
                    except ValueError:
                        try:
                            parsed_date = datetime.strptime(exit_date.strip(), '%d/%m/%Y')
                            exit_date_str = parsed_date.strftime('%Y-%m-%d')
                        except ValueError:
                            errors.append(f"Row {row_idx}: Invalid exit date format '{exit_date}'")
                            continue
                else:
                    errors.append(f"Row {row_idx}: Missing exit date")
                    continue

                # Validate date logic
                if entry_date_str >= exit_date_str:
                    errors.append(f"Row {row_idx}: Entry date must be before exit date")
                    continue

                trips.append({
                    'name': name,
                    'country': country_name,
                    'entry_date': entry_date_str,
                    'exit_date': exit_date_str,
                    'row': row_idx
                })

            except Exception as e:
                errors.append(f"Row {row_idx}: Error processing row - {str(e)}")

        return trips, errors

    except Exception as e:
        return [], [f"Error reading Excel file: {str(e)}"]

def find_employee_by_name(name):
    """Find employee by name (case-insensitive)"""
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT id, name FROM employees WHERE LOWER(name) = LOWER(?)', (name,))
    result = c.fetchone()
    conn.close()
    return result

def trip_exists(employee_id, entry_date, exit_date):
    """Check if trip already exists"""
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT id FROM trips WHERE employee_id = ? AND entry_date = ? AND exit_date = ?',
              (employee_id, entry_date, exit_date))
    result = c.fetchone()
    conn.close()
    return result is not None

@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/privacy')
def privacy():
    return render_template('privacy.html',
                         current_date=datetime.now().strftime('%Y-%m-%d'),
                         retention_months=CONFIG.get('RETENTION_MONTHS', 36),
                         session_timeout_minutes=CONFIG.get('SESSION_IDLE_TIMEOUT_MINUTES', 30),
                         max_login_attempts=MAX_ATTEMPTS,
                         rate_limit_window=ATTEMPT_WINDOW_SECONDS // 60)

@app.route('/help')
@login_required
def help_page():
    """Display help and tutorial page."""
    import json
    help_file = os.path.join(os.path.dirname(__file__), 'help_content.json')
    
    try:
        with open(help_file, 'r') as f:
            help_content = json.load(f)
    except Exception as e:
        # Fallback if help file doesn't exist
        help_content = {
            'page_title': 'Help & User Guide',
            'page_subtitle': 'Learn how to use the EU Trip Tracker',
            'sections': [],
            'tutorial_steps': [],
            'quick_tips': []
        }
    
    return render_template('help.html', help_content=help_content)

@app.route('/entry-requirements')
@login_required
def entry_requirements():
    """Display EU entry requirements for UK citizens."""
    return render_template('entry_requirements.html', countries=EU_ENTRY_DATA)

@app.route('/test-modal')
@login_required
def test_modal():
    """Test route to bypass caching issues."""
    return render_template('entry_requirements.html', countries=EU_ENTRY_DATA)

@app.route('/modal-debug')
def modal_debug():
    """Debug modal functionality."""
    return render_template('modal_debug.html')


@app.route('/api/entry-requirements')
@login_required
def entry_requirements_api():
    """API endpoint returning EU entry requirements as JSON."""
    return jsonify(EU_ENTRY_DATA)

@app.route('/api/entry-requirements/reload', methods=['POST'])
@login_required
def reload_entry_requirements():
    """Reload EU entry requirements data from JSON file."""
    global EU_ENTRY_DATA
    try:
        data_file_path = os.path.join(os.path.dirname(__file__), 'data', 'eu_entry_requirements.json')
        with open(data_file_path, 'r', encoding='utf-8') as f:
            EU_ENTRY_DATA = json.load(f)
        return jsonify({
            'success': True, 
            'message': f'Successfully reloaded {len(EU_ENTRY_DATA)} countries',
            'count': len(EU_ENTRY_DATA)
        })
    except Exception as e:
        return jsonify({
            'success': False, 
            'message': f'Error reloading data: {str(e)}'
        }), 500

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        ip = request.remote_addr or 'local'
        now = datetime.utcnow()
        attempts = LOGIN_ATTEMPTS.get(ip)
        if attempts:
            count, first_ts = attempts
            if (now - first_ts).total_seconds() <= ATTEMPT_WINDOW_SECONDS and count >= MAX_ATTEMPTS:
                write_audit(CONFIG['AUDIT_LOG_PATH'], 'login_rate_limited', 'admin', {'ip': ip})
                return render_template('login.html', error='Too many attempts. Try again later.')
            if (now - first_ts).total_seconds() > ATTEMPT_WINDOW_SECONDS:
                LOGIN_ATTEMPTS[ip] = (0, now)
        else:
            LOGIN_ATTEMPTS[ip] = (0, now)

        password = request.form.get('password', '')
        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT password_hash FROM admin WHERE id = 1')
        admin = c.fetchone()
        if not admin:
            conn.close()
            return render_template('login.html', error='Setup incomplete')
        ok, new_hash = verify_and_upgrade_password(admin['password_hash'], password)
        if ok:
            if new_hash:
                c.execute('UPDATE admin SET password_hash = ? WHERE id = 1', (new_hash,))
                conn.commit()
            conn.close()
            session['logged_in'] = True
            session['last_activity'] = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')
            write_audit(CONFIG['AUDIT_LOG_PATH'], 'login_success', 'admin', {'ip': ip})
            return redirect(url_for('dashboard'))
        # failed
        conn.close()
        count, first_ts = LOGIN_ATTEMPTS.get(ip, (0, now))
        LOGIN_ATTEMPTS[ip] = (count + 1, first_ts)
        write_audit(CONFIG['AUDIT_LOG_PATH'], 'login_failure', 'admin', {'ip': ip})
        return render_template('login.html', error='Invalid password')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    write_audit(CONFIG['AUDIT_LOG_PATH'], 'logout', 'admin', {})
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    # Get sort parameter from query string
    sort_by = request.args.get('sort', 'first_name')
    
    conn = get_db()
    c = conn.cursor()
    
    # Determine sort order based on parameter
    if sort_by == 'last_name':
        # Sort by last name (everything after the first space)
        c.execute('SELECT id, name FROM employees ORDER BY SUBSTR(name, INSTR(name, " ") + 1), name')
    else:
        # Default: sort by first name (full name)
        c.execute('SELECT id, name FROM employees ORDER BY name')
    
    employees = c.fetchall()
    employee_data = []
    at_risk_employees = []
    today = datetime.now().date()
    
    # Get risk thresholds from config
    risk_thresholds = CONFIG.get('RISK_THRESHOLDS', {'green': 30, 'amber': 10})

    for emp in employees:
        # Get all trips for this employee
        c.execute('SELECT entry_date, exit_date, country FROM trips WHERE employee_id = ?', (emp['id'],))
        trips = [{'entry_date': t['entry_date'], 'exit_date': t['exit_date'], 'country': t['country']} for t in c.fetchall()]
        
        # Calculate presence days and usage using new rolling90 module
        presence = presence_days(trips)
        days_used = days_used_in_window(presence, today)
        days_remaining = calculate_days_remaining(presence, today)
        risk_level = get_risk_level(days_remaining, risk_thresholds)
        
        # Calculate earliest safe entry date
        safe_entry = earliest_safe_entry(presence, today)
        
        # Calculate days until compliant for at-risk employees (days_remaining < 0)
        days_until_compliant_val = None
        compliance_date = None
        if days_remaining < 0:
            days_until_compliant_val, compliance_date = days_until_compliant(presence, today)

        # Get total trip count
        c.execute('SELECT COUNT(*) FROM trips WHERE employee_id = ?', (emp['id'],))
        trip_count = c.fetchone()[0]

        # Get next upcoming trip
        c.execute('SELECT country, entry_date FROM trips WHERE employee_id = ? AND entry_date > ? ORDER BY entry_date ASC LIMIT 1',
                 (emp['id'], today.strftime('%Y-%m-%d')))
        next_trip = c.fetchone()

        # Get recent trips for the card view
        c.execute('SELECT country, entry_date, exit_date FROM trips WHERE employee_id = ? ORDER BY exit_date DESC LIMIT 5', (emp['id'],))
        recent_trips = c.fetchall()

        emp_data = {
            'id': emp['id'],
            'name': emp['name'],
            'days_used': days_used,
            'days_remaining': days_remaining,
            'risk_level': risk_level,  # 'green', 'amber', or 'red'
            'status': 'danger' if days_used > 80 else ('warning' if days_used > 60 else 'safe'),
            'trip_count': trip_count,
            'next_trip': next_trip,
            'recent_trips': recent_trips,
            'earliest_safe_entry': safe_entry.strftime('%Y-%m-%d') if safe_entry else None,
            'days_until_compliant': days_until_compliant_val,
            'compliance_date': compliance_date.strftime('%Y-%m-%d') if compliance_date else None
        }

        # Separate at-risk employees (>76 days used) from safe employees
        # 76 days = 2 working weeks notice before 90-day limit
        if days_used > 76:
            at_risk_employees.append(emp_data)
        else:
            employee_data.append(emp_data)

    conn.close()
    return render_template('dashboard.html', 
                         employees=employee_data, 
                         at_risk_employees=at_risk_employees, 
                         current_sort=sort_by,
                         risk_thresholds=risk_thresholds)

@app.route('/employee/<int:employee_id>')
@login_required
def employee_detail(employee_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT name FROM employees WHERE id = ?', (employee_id,))
    employee = c.fetchone()
    if not employee:
        conn.close()
        return redirect(url_for('dashboard'))
    
    # Get all trips for rolling calculations
    c.execute('SELECT entry_date, exit_date, country FROM trips WHERE employee_id = ?', (employee_id,))
    all_trips = [{'entry_date': t['entry_date'], 'exit_date': t['exit_date'], 'country': t['country']} for t in c.fetchall()]
    
    # Calculate presence days and usage using new rolling90 module
    today = datetime.now().date()
    presence = presence_days(all_trips)
    days_used = days_used_in_window(presence, today)
    days_remaining = calculate_days_remaining(presence, today)
    
    # Calculate earliest safe entry date
    safe_entry = earliest_safe_entry(presence, today)
    
    # Get detailed trip list for display
    c.execute('SELECT id, country, entry_date, exit_date, travel_days, validation_note FROM trips WHERE employee_id = ? ORDER BY entry_date DESC', (employee_id,))
    trips = c.fetchall()

    # Calculate days_until_safe if days_used >= 90
    days_until_safe = None
    if days_used >= 90 and safe_entry:
        days_until_safe = (safe_entry - today).days

    # Process trips to add calculated fields
    processed_trips = []
    
    for trip in trips:
        entry_date = datetime.strptime(trip['entry_date'], '%Y-%m-%d').date()
        exit_date = datetime.strptime(trip['exit_date'], '%Y-%m-%d').date()
        
        # Calculate duration
        duration = (exit_date - entry_date).days + 1
        
        # Determine status
        if exit_date < today:
            status = 'completed'
        elif entry_date <= today <= exit_date:
            status = 'current'
        else:
            status = 'future'
        
        processed_trip = dict(trip)
        processed_trip['duration'] = duration
        processed_trip['status'] = status
        processed_trip['entry_date_obj'] = entry_date
        processed_trip['exit_date_obj'] = exit_date
        processed_trips.append(processed_trip)

    conn.close()
    
    # Get risk thresholds from config
    risk_thresholds = CONFIG.get('RISK_THRESHOLDS', {'green': 30, 'amber': 10})
    risk_level = get_risk_level(days_remaining, risk_thresholds)
    
    return render_template('employee_detail.html', 
                         employee={'id': employee_id, 'name': employee['name']}, 
                         trips=processed_trips, 
                         days_used=days_used,
                         days_remaining=days_remaining,
                         risk_level=risk_level,
                         days_until_safe=days_until_safe,
                         earliest_safe_entry=safe_entry,
                         today=today)

@app.route('/employee/<int:employee_id>/calendar')
@login_required
def employee_calendar(employee_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT name FROM employees WHERE id = ?', (employee_id,))
    employee = c.fetchone()
    if not employee:
        conn.close()
        return redirect(url_for('dashboard'))

    # Get all trips for the employee
    c.execute('SELECT id, country, entry_date, exit_date FROM trips WHERE employee_id = ? ORDER BY entry_date DESC', (employee_id,))
    trips = c.fetchall()

    # Calculate current 180-day window
    today = datetime.now().date()
    window_start = today - timedelta(days=180)
    window_end = today

    # Calculate days used within the 180-day window
    days_used = calculate_eu_days(employee_id)

    # Prepare trip data with additional info for calendar display
    calendar_trips = []
    for trip in trips:
        entry_date = datetime.strptime(trip['entry_date'], '%Y-%m-%d').date()
        exit_date = datetime.strptime(trip['exit_date'], '%Y-%m-%d').date()

        # Check if trip overlaps with 180-day window
        overlaps_window = exit_date >= window_start and entry_date <= window_end

        # Calculate days within window for this trip
        if overlaps_window:
            effective_entry = max(entry_date, window_start)
            effective_exit = min(exit_date, window_end)
            days_in_window = (effective_exit - effective_entry).days + 1
        else:
            days_in_window = 0

        calendar_trips.append({
            'id': trip['id'],
            'country': trip['country'],
            'entry_date': entry_date,
            'exit_date': exit_date,
            'overlaps_window': overlaps_window,
            'days_in_window': days_in_window,
            'trip_duration': (exit_date - entry_date).days + 1
        })

    # Calculate timeline range (last 12 months)
    timeline_start = today - timedelta(days=365)
    timeline_end = today + timedelta(days=30)  # Show a bit into the future

    conn.close()
    return render_template('employee_calendar.html',
                         employee={'id': employee_id, 'name': employee['name']},
                         trips=calendar_trips,
                         days_used=days_used,
                         window_start=window_start,
                         window_end=window_end,
                         timeline_start=timeline_start,
                         timeline_end=timeline_end,
                         today=today)

@app.route('/calendar/<int:employee_id>')
@login_required
def calendar(employee_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT name FROM employees WHERE id = ?', (employee_id,))
    employee = c.fetchone()
    if not employee:
        conn.close()
        return redirect(url_for('dashboard'))

    # Calculate current 180-day window
    today = datetime.now().date()
    window_start = today - timedelta(days=180)
    window_end = today

    # Get trips for the specified date range (6 months ago to 3 months future)
    timeline_start = today - timedelta(days=180)  # 6 months ago
    timeline_end = today + timedelta(days=90)     # 3 months future

    c.execute('SELECT id, country, entry_date, exit_date FROM trips WHERE employee_id = ? AND exit_date >= ? AND entry_date <= ? ORDER BY entry_date',
              (employee_id, timeline_start.strftime('%Y-%m-%d'), timeline_end.strftime('%Y-%m-%d')))
    trips = c.fetchall()

    # Calculate days used within the 180-day window
    days_used = calculate_eu_days(employee_id)

    # Prepare trip data with color coding
    calendar_trips = []
    for trip in trips:
        entry_date = datetime.strptime(trip['entry_date'], '%Y-%m-%d').date()
        exit_date = datetime.strptime(trip['exit_date'], '%Y-%m-%d').date()

        # Determine trip type for color coding
        if exit_date < today:
            trip_type = 'past'
        elif entry_date <= today <= exit_date:
            trip_type = 'current'
        else:
            trip_type = 'future'

        # Check if trip overlaps with 180-day window
        overlaps_window = exit_date >= window_start and entry_date <= window_end

        # Calculate days within window for this trip
        if overlaps_window:
            effective_entry = max(entry_date, window_start)
            effective_exit = min(exit_date, window_end)
            days_in_window = (effective_exit - effective_entry).days + 1
        else:
            days_in_window = 0

        calendar_trips.append({
            'id': trip['id'],
            'country': trip['country'],
            'entry_date': entry_date,
            'exit_date': exit_date,
            'trip_type': trip_type,
            'overlaps_window': overlaps_window,
            'days_in_window': days_in_window,
            'trip_duration': (exit_date - entry_date).days + 1
        })

    conn.close()
    return render_template('calendar.html',
                         employee={'id': employee_id, 'name': employee['name']},
                         trips=calendar_trips,
                         days_used=days_used,
                         window_start=window_start,
                         window_end=window_end,
                         timeline_start=timeline_start,
                         timeline_end=timeline_end,
                         today=today)

def calculate_forecast_eu_days(employee_id, forecast_entry_date, forecast_exit_date):
    """Calculate Schengen days including a forecast trip without saving to database.
    Ireland trips are excluded from calculations."""
    conn = get_db()
    c = conn.cursor()
    today = datetime.now().date()

    # Get all existing trips for this employee (including country info)
    c.execute('SELECT entry_date, exit_date, country FROM trips WHERE employee_id = ?', (employee_id,))
    existing_trips = [{'entry_date': t['entry_date'], 'exit_date': t['exit_date'], 'country': t['country']} for t in c.fetchall()]
    conn.close()

    # Calculate current days used using rolling90 logic (filters Ireland)
    current_presence = presence_days(existing_trips)
    current_days_used = days_used_in_window(current_presence, today)

    # Calculate forecast trip dates
    forecast_entry = datetime.strptime(forecast_entry_date, '%Y-%m-%d').date()
    forecast_exit = datetime.strptime(forecast_exit_date, '%Y-%m-%d').date()

    # Add the forecast trip to the trips list for calculation
    # Note: We assume the forecast trip is to a Schengen country (not Ireland)
    # If needed, this could be made configurable
    all_trips = existing_trips + [{
        'entry_date': forecast_entry_date,
        'exit_date': forecast_exit_date,
        'country': 'FR'  # Default to France (Schengen) - could be made configurable
    }]

    # Calculate forecast days using rolling90 logic
    forecast_presence = presence_days(all_trips)
    forecast_total_days = days_used_in_window(forecast_presence, forecast_exit)

    return {
        'current_days_used': current_days_used,
        'forecast_total_days': forecast_total_days,
        'forecast_trip_days': (forecast_exit - forecast_entry).days + 1
    }

@app.route('/forecast-trip', methods=['POST'])
@login_required
def forecast_trip():
    try:
        data = request.get_json()
        employee_id = int(data.get('employee_id'))
        entry_date = data.get('entry_date')
        exit_date = data.get('exit_date')
        country = data.get('country', '')

        # Validate inputs
        if not all([employee_id, entry_date, exit_date]):
            return jsonify({
                'error': 'Missing required fields: employee_id, entry_date, exit_date'
            }), 400

        # Validate date format and logic
        try:
            entry_dt = datetime.strptime(entry_date, '%Y-%m-%d').date()
            exit_dt = datetime.strptime(exit_date, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

        if exit_dt < entry_dt:
            return jsonify({'error': 'Exit date cannot be before entry date'}), 400

        # Check if employee exists
        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT name FROM employees WHERE id = ?', (employee_id,))
        employee = c.fetchone()
        conn.close()

        if not employee:
            return jsonify({'error': 'Employee not found'}), 404

        # Calculate forecast
        current_days_used = calculate_eu_days(employee_id)
        forecast_data = calculate_forecast_eu_days(employee_id, entry_date, exit_date)

        total_days_after = forecast_data['forecast_total_days']
        trip_days = forecast_data['forecast_trip_days']

        # Determine status
        if total_days_after > 90:
            status = 'over_limit'
            message = f'This trip would exceed the 90-day limit by {total_days_after - 90} days.'
        elif total_days_after >= 85:
            status = 'warning'
            message = f'This trip brings you close to the limit. Only {90 - total_days_after} days remaining.'
        else:
            status = 'valid'
            message = f'This trip is valid. You\'ll have {90 - total_days_after} days remaining.'

        return jsonify({
            'employee_name': employee['name'],
            'current_days_used': current_days_used,
            'planned_trip_days': trip_days,
            'total_days_after': total_days_after,
            'status': status,
            'message': message,
            'country': country,
            'entry_date': entry_date,
            'exit_date': exit_date,
            'remaining_days': 90 - total_days_after
        })

    except Exception as e:
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@app.route('/add_employee', methods=['POST'])
@login_required
def add_employee():
    name = request.form.get('name', '').strip()
    
    # Validate employee name
    validated_name, error = validate_employee_name(name)
    if error:
        flash(f'Error: {error}')
        return redirect(url_for('dashboard'))
    
    conn = get_db()
    try:
        conn.execute('INSERT INTO employees (name) VALUES (?)', (validated_name,))
        conn.commit()
        write_audit(CONFIG['AUDIT_LOG_PATH'], 'employee_added', 'admin', {'name': validated_name})
        flash(f'Successfully added employee: {validated_name}')
    except sqlite3.IntegrityError:
        flash(f'Error: Employee "{validated_name}" already exists')
    except Exception as e:
        flash(f'Error adding employee: {str(e)}')
    finally:
        conn.close()
    
    return redirect(url_for('dashboard'))

@app.route('/add_trip', methods=['POST'])
@login_required
def add_trip():
    employee_id = request.form.get('employee_id')
    country_code_input = request.form.get('country_code', '').strip()
    entry_date_str = request.form.get('entry_date', '').strip()
    exit_date_str = request.form.get('exit_date', '').strip()
    override_validation = request.form.get('override_validation') == 'true'
    validation_note = request.form.get('validation_note', '').strip()
    
    # Validate inputs
    if not employee_id:
        flash('Error: Employee ID is required')
        return redirect(url_for('dashboard'))
    
    # Validate country code
    country_code, error = validate_country_code(country_code_input)
    if error:
        flash(f'Error: {error}')
        return redirect(url_for('employee_detail', employee_id=employee_id))
    
    # Validate dates
    entry_date, error = validate_date(entry_date_str, "Entry date")
    if error:
        flash(f'Error: {error}')
        return redirect(url_for('employee_detail', employee_id=employee_id))
    
    exit_date, error = validate_date(exit_date_str, "Exit date")
    if error:
        flash(f'Error: {error}')
        return redirect(url_for('employee_detail', employee_id=employee_id))
    
    # Get existing trips for this employee
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT id, entry_date, exit_date FROM trips WHERE employee_id = ?', (employee_id,))
    existing_trips = [{'id': t['id'], 'entry_date': t['entry_date'], 'exit_date': t['exit_date']} for t in c.fetchall()]
    
    # Validate trip using trip_validator
    hard_errors, soft_warnings = validate_trip(existing_trips, entry_date, exit_date)
    
    # Block if there are hard errors
    if hard_errors:
        for error in hard_errors:
            flash(f'Error: {error}', 'error')
        conn.close()
        return redirect(url_for('employee_detail', employee_id=employee_id))
    
    # Show warnings but allow override
    if soft_warnings and not override_validation:
        for warning in soft_warnings:
            flash(f'Warning: {warning}', 'warning')
        # Store trip data in session for potential override
        session['pending_trip'] = {
            'employee_id': employee_id,
            'country_code': country_code,
            'entry_date': entry_date.strftime('%Y-%m-%d'),
            'exit_date': exit_date.strftime('%Y-%m-%d'),
            'warnings': soft_warnings
        }
        conn.close()
        return redirect(url_for('employee_detail', employee_id=employee_id) + '#trip-warnings')
    
    try:
        # Insert trip with validation note if overridden
        validation_note_text = validation_note if (soft_warnings and override_validation) else None
        
        conn.execute('INSERT INTO trips (employee_id, country, entry_date, exit_date, validation_note) VALUES (?, ?, ?, ?, ?)', 
                    (employee_id, country_code, entry_date.strftime('%Y-%m-%d'), exit_date.strftime('%Y-%m-%d'), validation_note_text))
        conn.commit()
        conn.close()
        
        # Clear pending trip from session
        session.pop('pending_trip', None)
        
        write_audit(CONFIG['AUDIT_LOG_PATH'], 'trip_added', 'admin', {
            'employee_id': employee_id,
            'country': country_code,
            'entry_date': entry_date.strftime('%Y-%m-%d'),
            'exit_date': exit_date.strftime('%Y-%m-%d'),
            'validation_overridden': override_validation,
            'validation_note': validation_note_text
        })
        
        flash('Trip added successfully', 'success')
    except Exception as e:
        flash(f'Error adding trip: {str(e)}', 'error')
    
    return redirect(url_for('employee_detail', employee_id=employee_id))

@app.route('/delete_trip/<int:trip_id>', methods=['POST'])
@login_required
def delete_trip(trip_id):
    employee_id = request.form.get('employee_id')
    conn = get_db()
    conn.execute('DELETE FROM trips WHERE id = ?', (trip_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('employee_detail', employee_id=employee_id))

@app.route('/delete_employee/<int:employee_id>', methods=['POST'])
@login_required
def delete_employee(employee_id):
    conn = get_db()
    conn.execute('DELETE FROM employees WHERE id = ?', (employee_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('dashboard'))

@app.route('/change_password', methods=['POST'])
@login_required
def change_password():
    new_password = request.form.get('new_password', '')
    
    # Validate password strength
    if len(new_password) < 8:
        flash('Error: Password must be at least 8 characters long')
        return redirect(url_for('dashboard'))
    
    # Check for password complexity (at least one letter and one number)
    if not re.search(r'[A-Za-z]', new_password) or not re.search(r'[0-9]', new_password):
        flash('Error: Password must contain both letters and numbers')
        return redirect(url_for('dashboard'))
    
    try:
        conn = get_db()
        new_hash = hasher.hash(new_password)
        conn.execute('UPDATE admin SET password_hash = ? WHERE id = 1', (new_hash,))
        conn.commit()
        conn.close()
        
        write_audit(CONFIG['AUDIT_LOG_PATH'], 'password_changed', 'admin', {})
        flash('Password changed successfully')
    except Exception as e:
        flash(f'Error changing password: {str(e)}')
    
    return redirect(url_for('dashboard'))

@app.route('/bulk_add_trip', methods=['GET', 'POST'])
@login_required
def bulk_add_trip():
    if request.method == 'POST':
        employee_ids = request.form.getlist('employee_ids')
        country_code = (request.form.get('country_code', '') or '').strip().upper()
        country_code = re.sub(r'[^A-Z]', '', country_code)
        entry_date = request.form.get('entry_date', '')
        exit_date = request.form.get('exit_date', '')

        # Validation
        if not employee_ids:
            conn = get_db()
            c = conn.cursor()
            c.execute('SELECT id, name FROM employees ORDER BY name')
            employees = c.fetchall()
            conn.close()
            return render_template('bulk_add_trip.html', employees=employees, error='Please select at least one employee.')

        if not country_code or not entry_date or not exit_date:
            conn = get_db()
            c = conn.cursor()
            c.execute('SELECT id, name FROM employees ORDER BY name')
            employees = c.fetchall()
            conn.close()
            return render_template('bulk_add_trip.html', employees=employees, error='Please fill in all fields.')

        try:
            if datetime.strptime(exit_date, '%Y-%m-%d') < datetime.strptime(entry_date, '%Y-%m-%d'):
                conn = get_db()
                c = conn.cursor()
                c.execute('SELECT id, name FROM employees ORDER BY name')
                employees = c.fetchall()
                conn.close()
                return render_template('bulk_add_trip.html', employees=employees, error='Exit date cannot be before entry date.')
        except ValueError:
            conn = get_db()
            c = conn.cursor()
            c.execute('SELECT id, name FROM employees ORDER BY name')
            employees = c.fetchall()
            conn.close()
            return render_template('bulk_add_trip.html', employees=employees, error='Invalid date format.')

        # Insert trips for selected employees
        conn = get_db()
        success_count = 0
        for employee_id in employee_ids:
            try:
                if country_code and COUNTRY_CODE_MAPPING.get(country_code):
                    conn.execute('INSERT INTO trips (employee_id, country, entry_date, exit_date) VALUES (?, ?, ?, ?)',
                               (employee_id, country_code, entry_date, exit_date))
                    success_count += 1
                else:
                    continue
            except:
                continue

        conn.commit()
        conn.close()

        if success_count > 0:
            message = f'Successfully added trip to {success_count} employee{"s" if success_count != 1 else ""}.'
            return redirect(url_for('dashboard'))
        else:
            conn = get_db()
            c = conn.cursor()
            c.execute('SELECT id, name FROM employees ORDER BY name')
            employees = c.fetchall()
            conn.close()
            return render_template('bulk_add_trip.html', employees=employees, error='Failed to add trips. Please try again.')

    # GET request - show the form
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT id, name FROM employees ORDER BY name')
    employees = c.fetchall()
    conn.close()
    return render_template('bulk_add_trip.html', employees=employees)

@app.route('/import_excel', methods=['GET', 'POST'])
@login_required
def import_excel():
    if request.method == 'POST':
        # Handle file upload
        if 'excel_file' not in request.files:
            flash('No file uploaded')
            return redirect(request.url)

        file = request.files['excel_file']
        if file.filename == '':
            flash('No file selected')
            return redirect(request.url)

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)

            try:
                # Use importer.process_excel to parse and save, returning a summary
                from importer import process_excel
                summary = process_excel(filepath, db_path=DATABASE)

                # Clean up uploaded file
                os.remove(filepath)

                # Flash basic message
                if summary.get('trips_added', 0) > 0:
                    flash(f"Imported {summary['trips_added']} trips. {summary.get('duplicates_skipped', 0)} duplicates skipped. {summary.get('uk_jobs_skipped', 0)} UK jobs ignored.")
                else:
                    flash('No valid trips found in the uploaded file.')

                # Render page with detailed summary
                return render_template('import_excel.html', import_summary=summary)

            except Exception as e:
                # Clean up uploaded file
                os.remove(filepath)
                flash(f'Error importing Excel file: {str(e)}')
                print(f"Import error: {e}")
                return redirect(request.url)
        else:
            flash('Invalid file format. Please upload an Excel file (.xlsx or .xls)')
            return redirect(request.url)

    return render_template('import_excel.html')

@app.route('/import_preview')
@login_required
def import_preview():
    if 'import_preview' not in session:
        flash('No import data found')
        return redirect(url_for('import_excel'))

    data = session['import_preview']
    return render_template('import_preview.html',
                         valid_trips=data['valid_trips'],
                         errors=data['errors'])

@app.route('/confirm_import', methods=['POST'])
@login_required
def confirm_import():
    if 'import_preview' not in session:
        flash('No import data found')
        return redirect(url_for('import_excel'))

    valid_trips = session['import_preview']['valid_trips']

    # Import the trips
    conn = get_db()
    c = conn.cursor()
    success_count = 0

    for trip in valid_trips:
        try:
            c.execute('INSERT INTO trips (employee_id, country, entry_date, exit_date) VALUES (?, ?, ?, ?)',
                     (trip['employee_id'], trip['country'], trip['entry_date'], trip['exit_date']))
            success_count += 1
        except Exception as e:
            print(f"Error importing trip: {e}")

    conn.commit()
    conn.close()

    # Clear session data
    session.pop('import_preview', None)

    flash(f'Successfully imported {success_count} trips')
    return redirect(url_for('dashboard'))

@app.route('/cancel_import', methods=['POST'])
@login_required
def cancel_import():
    session.pop('import_preview', None)
    flash('Import cancelled')
    return redirect(url_for('import_excel'))

@app.route('/clear_imported_data', methods=['POST'])
@login_required
def clear_imported_data():
    """Clear all imported trip data - secure implementation"""
    try:
        conn = get_db()
        c = conn.cursor()
        
        # Count records before deletion for audit
        c.execute('SELECT COUNT(*) FROM trips')
        trips_count = c.fetchone()[0]
        c.execute('SELECT COUNT(*) FROM employees')
        employees_count = c.fetchone()[0]
        
        # Delete all data
        c.execute('DELETE FROM trips')
        c.execute('DELETE FROM employees')
        conn.commit()
        conn.close()
        
        # Log the action
        write_audit(CONFIG['AUDIT_LOG_PATH'], 'clear_imported_data', 'admin', {
            'trips_deleted': trips_count,
            'employees_deleted': employees_count,
            'timestamp': datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')
        })
        
        flash(f'Successfully cleared all imported data. {employees_count} employees and {trips_count} trips removed.')
        
    except Exception as e:
        flash(f'Error clearing data: {str(e)}')
    
    return redirect(url_for('import_excel'))

# REMOVED: Testing utility route - unsafe for production
# Use admin privacy tools for data management instead

# -------------------------
# DSAR (Data Subject Access Request) endpoints
# -------------------------

@app.route('/api/dsar/export')
@login_required
def dsar_export():
    """Export personal data for a given employee (by name or id)."""
    employee_name = request.args.get('employee_name', '').strip()
    employee_id = request.args.get('employee_id', '').strip()

    conn = get_db()
    c = conn.cursor()

    if employee_id:
        c.execute('SELECT id, name FROM employees WHERE id = ?', (employee_id,))
    elif employee_name:
        c.execute('SELECT id, name FROM employees WHERE LOWER(name) = LOWER(?)', (employee_name,))
    else:
        conn.close()
        return jsonify({'error': 'Provide employee_name or employee_id'}), 400

    emp = c.fetchone()
    if not emp:
        conn.close()
        return jsonify({'error': 'Employee not found'}), 404

    c.execute('SELECT country, entry_date, exit_date, travel_days FROM trips WHERE employee_id = ? ORDER BY entry_date ASC', (emp[0],))
    trips = [
        {
            'country': row[0],
            'entry_date': row[1],
            'exit_date': row[2],
            'travel_days': row[3] if row[3] is not None else 0
        }
        for row in c.fetchall()
    ]
    conn.close()

    export_payload = {
        'employee': {
            'id': emp[0],
            'name': emp[1]
        },
        'trips': trips,
        'generated_at': datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')
    }

    response = make_response(jsonify(export_payload))
    response.headers['Content-Disposition'] = f'attachment; filename=eu-trip-tracker-{emp[1]}-export.json'
    return response


@app.route('/api/dsar/delete', methods=['POST'])
@login_required
def dsar_delete():
    """Delete/anonymize personal data for a given employee (by id or name)."""
    try:
        data = request.get_json() or {}
    except Exception:
        data = {}

    employee_id = str(data.get('employee_id', '')).strip()
    employee_name = str(data.get('employee_name', '')).strip()

    conn = get_db()
    c = conn.cursor()

    # Resolve employee
    if employee_id:
        c.execute('SELECT id, name FROM employees WHERE id = ?', (employee_id,))
    elif employee_name:
        c.execute('SELECT id, name FROM employees WHERE LOWER(name) = LOWER(?)', (employee_name,))
    else:
        conn.close()
        return jsonify({'error': 'Provide employee_name or employee_id'}), 400

    emp = c.fetchone()
    if not emp:
        conn.close()
        return jsonify({'error': 'Employee not found'}), 404

    # Delete trips and employee record
    c.execute('DELETE FROM trips WHERE employee_id = ?', (emp[0],))
    c.execute('DELETE FROM employees WHERE id = ?', (emp[0],))
    conn.commit()
    conn.close()

    return jsonify({'status': 'deleted', 'employee_id': emp[0], 'employee_name': emp[1]})

@app.route('/calendar_view')
@login_required
def calendar_view():
    # Render global calendar view with employee filter
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT id, name FROM employees ORDER BY name')
    employees = c.fetchall()
    conn.close()
    today = datetime.now().date()
    return render_template('global_calendar.html', employees=employees, today=today)

@app.route('/api/calendar_data')
@login_required
def api_calendar_data():
    try:
        # Parse optional params from FullCalendar
        start_param = request.args.get('start')
        end_param = request.args.get('end')
        employee_id_param = request.args.get('employee_id')

        today = datetime.now().date()
        allowed_start = today - timedelta(days=180)
        allowed_end = today + timedelta(days=56)

        # FullCalendar passes ISO strings (YYYY-MM-DD)
        def parse_date_or_default(value, default_value):
            try:
                if value:
                    return datetime.strptime(value[:10], '%Y-%m-%d').date()
            except Exception:
                pass
            return default_value

        requested_start = parse_date_or_default(start_param, allowed_start)
        requested_end = parse_date_or_default(end_param, allowed_end)

        # Clamp to allowed window
        query_start = max(requested_start, allowed_start)
        query_end = min(requested_end, allowed_end)

        conn = get_db()
        conn.row_factory = sqlite3.Row
        c = conn.cursor()

        # For accurate 180-day calculations near the start boundary, prefetch from (allowed_start - 180)
        prefetch_start = allowed_start - timedelta(days=180)

        sql = (
            'SELECT trips.id as trip_id, trips.employee_id, employees.name as employee_name, '
            'trips.country, trips.entry_date, trips.exit_date '
            'FROM trips JOIN employees ON trips.employee_id = employees.id '
            'WHERE trips.exit_date >= ? AND trips.entry_date <= ?'
        )
        params = [query_start.strftime('%Y-%m-%d'), query_end.strftime('%Y-%m-%d')]

        if employee_id_param:
            sql += ' AND trips.employee_id = ?'
            params.append(employee_id_param)

        sql += ' ORDER BY trips.entry_date ASC'

        c.execute(sql, params)
        rows = c.fetchall()

        # Prefetch trips for 180-day calculations per employee
        employee_ids = set([row['employee_id'] for row in rows])
        trips_by_employee = {}
        if employee_ids:
            if employee_id_param:
                prefetch_ids = [int(employee_id_param)]
            else:
                prefetch_ids = list(employee_ids)
            placeholders = ','.join(['?'] * len(prefetch_ids))
            pre_sql = (
                f'SELECT employee_id, entry_date, exit_date FROM trips '
                f'WHERE employee_id IN ({placeholders}) AND exit_date >= ? AND entry_date <= ?'
            )
            pre_params = prefetch_ids + [prefetch_start.strftime('%Y-%m-%d'), allowed_end.strftime('%Y-%m-%d')]
            c.execute(pre_sql, pre_params)
            pre_rows = c.fetchall()
            for r in pre_rows:
                trips_by_employee.setdefault(r['employee_id'], []).append({'entry_date': r['entry_date'], 'exit_date': r['exit_date']})

        events = []
        GREEN = '#10b981'
        YELLOW = '#f59e0b'
        RED = '#ef4444'

        for row in rows:
            emp_id = row['employee_id']
            emp_name = row['employee_name']
            country = row['country']
            entry_dt = datetime.strptime(row['entry_date'], '%Y-%m-%d').date()
            exit_dt = datetime.strptime(row['exit_date'], '%Y-%m-%d').date()

            # Duration inclusive
            duration_days = (exit_dt - entry_dt).days + 1

            # Current usage (as of today)
            emp_trips = trips_by_employee.get(emp_id, [])
            current_used = calculate_eu_days_at_date(emp_trips, today)

            # Forecast usage at the end of this trip
            forecast_used = calculate_eu_days_at_date(emp_trips, exit_dt)

            # Color rules
            if forecast_used >= 90:
                color = RED
                status = 'over_limit'
            elif forecast_used >= 80:
                color = YELLOW
                status = 'warning'
            else:
                color = GREEN
                status = 'ok'

            # FullCalendar's end is exclusive; add one day
            end_exclusive = (exit_dt + timedelta(days=1)).strftime('%Y-%m-%d')

            tooltip = (
                f"Employee: {emp_name}\n"
                f"Country: {country}\n"
                f"Trip: {row['entry_date']} → {row['exit_date']}\n"
                f"Duration: {duration_days} days\n"
                f"Used now: {current_used}/90\n"
                f"Used at trip end: {forecast_used}/90"
            )

            events.append({
                'id': f"{row['trip_id']}",
                'title': f"{emp_name} - {country}",
                'start': row['entry_date'],
                'end': end_exclusive,
                'allDay': True,
                'color': color,
                'extendedProps': {
                    'employeeId': emp_id,
                    'employeeName': emp_name,
                    'country': country,
                    'durationDays': duration_days,
                    'currentUsed': current_used,
                    'forecastUsed': forecast_used,
                    'status': status,
                    'tooltip': tooltip
                }
            })

        conn.close()
        return jsonify(events)
    except Exception as e:
        return jsonify({'error': f'Server error: {str(e)}'}), 500

# -------------------------
# GDPR Compliance Endpoints
# -------------------------

@app.route('/admin/privacy-tools')
@login_required
def admin_privacy_tools():
    """Privacy tools page for DSAR operations."""
    try:
        expired_trips = get_expired_trips(DATABASE, CONFIG.get('RETENTION_MONTHS', 36))
        
        # Get last purge timestamp from a simple file
        last_purge_file = os.path.join(CONFIG.get('DSAR_EXPORT_DIR', './exports'), '.last_purge')
        last_purge = None
        if os.path.exists(last_purge_file):
            try:
                with open(last_purge_file, 'r') as f:
                    last_purge = f.read().strip()
            except:
                pass
        
        return render_template('admin_privacy_tools.html',
                             retention_months=CONFIG.get('RETENTION_MONTHS', 36),
                             expired_count=len(expired_trips),
                             last_purge=last_purge)
    except Exception as e:
        flash(f'Error loading privacy tools: {str(e)}')
        return redirect(url_for('dashboard'))

@app.route('/admin/settings', methods=['GET', 'POST'])
@login_required
def admin_settings():
    """Admin settings page."""
    if request.method == 'POST':
        try:
            updates = {
                'RETENTION_MONTHS': int(request.form.get('retention_months', 36)),
                'SESSION_IDLE_TIMEOUT_MINUTES': int(request.form.get('session_timeout', 30)),
                'PASSWORD_HASH_SCHEME': request.form.get('password_scheme', 'argon2'),
                'DSAR_EXPORT_DIR': request.form.get('export_dir', './exports'),
                'AUDIT_LOG_PATH': request.form.get('audit_log', './logs/audit.log'),
                'SESSION_COOKIE_SECURE': 'session_cookie_secure' in request.form
            }
            
            global CONFIG
            CONFIG = save_config(updates)
            
            write_audit(CONFIG['AUDIT_LOG_PATH'], 'settings_updated', 'admin', updates)
            flash('Settings updated successfully')
            return redirect(url_for('admin_settings'))
        except Exception as e:
            flash(f'Error updating settings: {str(e)}')
    
    return render_template('admin_settings.html', config=CONFIG, message=request.args.get('message'), error=request.args.get('error'))

@app.route('/api/employees/search')
@login_required
def api_employees_search():
    """Search employees by name."""
    try:
        query = request.args.get('q', '').strip()
        if not query:
            return jsonify({'error': 'No search query provided'}), 400
        
        conn = get_db()
        c = conn.cursor()
        c.execute('''
            SELECT e.id, e.name, COUNT(t.id) as trip_count
            FROM employees e
            LEFT JOIN trips t ON e.id = t.employee_id
            WHERE LOWER(e.name) LIKE LOWER(?)
            GROUP BY e.id, e.name
            ORDER BY e.name
            LIMIT 10
        ''', (f'%{query}%',))
        
        employees = []
        for row in c.fetchall():
            employees.append({
                'id': row['id'],
                'name': row['name'],
                'trip_count': row['trip_count']
            })
        
        conn.close()
        return jsonify({'employees': employees})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/admin/dsar/export/<int:employee_id>')
@login_required
def admin_dsar_export(employee_id):
    """Generate and download DSAR export ZIP for employee."""
    try:
        result = create_dsar_export(
            DATABASE, 
            employee_id, 
            CONFIG.get('DSAR_EXPORT_DIR', './exports'),
            CONFIG.get('RETENTION_MONTHS', 36)
        )
        
        if not result.get('success'):
            flash(result.get('error', 'Unknown error'))
            return redirect(url_for('admin_privacy_tools'))
        
        write_audit(CONFIG['AUDIT_LOG_PATH'], 'dsar_export', 'admin', {
            'employee_id': employee_id,
            'employee_name': result['employee_name']
        })
        
        return send_file(
            result['file_path'],
            as_attachment=True,
            download_name=result['filename'],
            mimetype='application/zip'
        )
    except Exception as e:
        flash(f'Error exporting data: {str(e)}')
        return redirect(url_for('admin_privacy_tools'))

@app.route('/admin/dsar/delete/<int:employee_id>', methods=['POST'])
@login_required
def admin_dsar_delete(employee_id):
    """Delete employee and all associated data."""
    try:
        result = delete_employee_data(DATABASE, employee_id)
        
        if not result.get('success'):
            return jsonify(result), 404
        
        write_audit(CONFIG['AUDIT_LOG_PATH'], 'dsar_delete', 'admin', {
            'employee_id': employee_id,
            'employee_name': result['employee_name'],
            'trips_deleted': result['trips_deleted']
        })
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/dsar/rectify/<int:employee_id>', methods=['POST'])
@login_required
def admin_dsar_rectify(employee_id):
    """Rectify (update) employee name."""
    try:
        data = request.get_json() or {}
        new_name = data.get('new_name', '').strip()
        
        result = rectify_employee_name(DATABASE, employee_id, new_name)
        
        if not result.get('success'):
            return jsonify(result), 400
        
        write_audit(CONFIG['AUDIT_LOG_PATH'], 'dsar_rectify', 'admin', {
            'employee_id': employee_id,
            'old_name': result['old_name'],
            'new_name': result['new_name']
        })
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/dsar/anonymize/<int:employee_id>', methods=['POST'])
@login_required
def admin_dsar_anonymize(employee_id):
    """Anonymize employee name."""
    try:
        result = anonymize_employee(DATABASE, employee_id)
        
        if not result.get('success'):
            return jsonify(result), 404
        
        write_audit(CONFIG['AUDIT_LOG_PATH'], 'dsar_anonymize', 'admin', {
            'employee_id': employee_id,
            'old_name': result['old_name'],
            'new_name': result['new_name']
        })
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/retention/purge', methods=['POST'])
@login_required
def admin_retention_purge():
    """Run retention purge to delete old trips."""
    try:
        result = purge_expired_trips(DATABASE, CONFIG.get('RETENTION_MONTHS', 36))
        
        # Save last purge timestamp
        last_purge_file = os.path.join(CONFIG.get('DSAR_EXPORT_DIR', './exports'), '.last_purge')
        os.makedirs(os.path.dirname(last_purge_file), exist_ok=True)
        with open(last_purge_file, 'w') as f:
            f.write(datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC'))
        
        write_audit(CONFIG['AUDIT_LOG_PATH'], 'retention_purge', 'admin', result)
        
        return jsonify({
            'success': True,
            'trips_deleted': result['trips_deleted'],
            'employees_deleted': result['employees_deleted'],
            'cutoff_date': result['cutoff_date']
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/retention/preview')
@login_required
def api_retention_preview():
    """Preview trips that would be purged."""
    try:
        expired_trips = get_expired_trips(DATABASE, CONFIG.get('RETENTION_MONTHS', 36))
        
        # Count unique employees
        employee_ids = set([trip['employee_id'] for trip in expired_trips])
        
        return jsonify({
            'trips_count': len(expired_trips),
            'employees_affected': len(employee_ids)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/admin/retention/expired')
@login_required
def admin_retention_expired():
    """Show list of expired trips."""
    try:
        expired_trips = get_expired_trips(DATABASE, CONFIG.get('RETENTION_MONTHS', 36))
        return render_template('expired_trips.html', 
                             trips=expired_trips,
                             retention_months=CONFIG.get('RETENTION_MONTHS', 36))
    except Exception as e:
        flash(f'Error loading expired trips: {str(e)}')
        return redirect(url_for('admin_privacy_tools'))

@app.route('/export/trips/csv')
@login_required
def export_trips_csv_endpoint():
    """Export all trips as CSV."""
    try:
        employee_id = request.args.get('employee_id', type=int)
        csv_data = export_trips_csv(DATABASE, employee_id)
        
        response = make_response(csv_data)
        filename = f"trips_employee_{employee_id}.csv" if employee_id else "trips_all.csv"
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'text/csv'
        
        write_audit(CONFIG['AUDIT_LOG_PATH'], 'export_csv', 'admin', {'employee_id': employee_id})
        return response
    except Exception as e:
        flash(f'Error exporting CSV: {str(e)}')
        return redirect(url_for('dashboard'))

@app.route('/export/employee/<int:employee_id>/pdf')
@login_required
def export_employee_pdf(employee_id):
    """Export employee report as PDF."""
    try:
        pdf_bytes = export_employee_report_pdf(DATABASE, employee_id)
        
        # Get employee name for filename
        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT name FROM employees WHERE id = ?', (employee_id,))
        emp = c.fetchone()
        conn.close()
        
        filename = f"report_{emp['name'].replace(' ', '_')}.pdf" if emp else f"report_{employee_id}.pdf"
        
        response = make_response(pdf_bytes)
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'application/pdf'
        
        write_audit(CONFIG['AUDIT_LOG_PATH'], 'export_pdf', 'admin', {'employee_id': employee_id})
        return response
    except Exception as e:
        flash(f'Error exporting PDF: {str(e)}')
        return redirect(url_for('employee_detail', employee_id=employee_id))

@app.route('/export/all/pdf')
@login_required
def export_all_pdf():
    """Export all employees summary as PDF."""
    try:
        pdf_bytes = export_all_employees_report_pdf(DATABASE)
        
        response = make_response(pdf_bytes)
        response.headers['Content-Disposition'] = 'attachment; filename=compliance_summary.pdf'
        response.headers['Content-Type'] = 'application/pdf'
        
        write_audit(CONFIG['AUDIT_LOG_PATH'], 'export_all_pdf', 'admin', {})
        return response
    except Exception as e:
        flash(f'Error exporting PDF: {str(e)}')
        return redirect(url_for('dashboard'))

@app.route('/admin/delete-all-data', methods=['POST'])
@login_required
def delete_all_data():
    """Delete ALL data from the system - GDPR compliance tool"""
    try:
        # Verify confirmation
        confirmation = request.form.get('confirmation', '').strip()
        if confirmation != 'DELETE ALL DATA':
            flash('Error: Incorrect confirmation phrase. Data was not deleted.')
            return redirect(url_for('admin_settings'))
        
        conn = get_db()
        c = conn.cursor()
        
        # Count records before deletion
        c.execute('SELECT COUNT(*) FROM trips')
        trips_count = c.fetchone()[0]
        c.execute('SELECT COUNT(*) FROM employees')
        employees_count = c.fetchone()[0]
        
        # Delete all data
        c.execute('DELETE FROM trips')
        c.execute('DELETE FROM employees')
        conn.commit()
        conn.close()
        
        # Log the action
        write_audit(CONFIG['AUDIT_LOG_PATH'], 'delete_all_data', 'admin', {
            'trips_deleted': trips_count,
            'employees_deleted': employees_count,
            'timestamp': datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')
        })
        
        # Also delete exported files
        try:
            export_dir = CONFIG.get('DSAR_EXPORT_DIR', './exports')
            if os.path.exists(export_dir):
                for filename in os.listdir(export_dir):
                    if filename.endswith(('.zip', '.json', '.csv', '.pdf')):
                        os.remove(os.path.join(export_dir, filename))
        except Exception as e:
            print(f"Warning: Could not delete export files: {e}")
        
        flash(f'All data deleted successfully. {employees_count} employees and {trips_count} trips removed.')
        return redirect(url_for('admin_settings'))
        
    except Exception as e:
        flash(f'Error deleting data: {str(e)}')
        return redirect(url_for('admin_settings'))

if __name__ == '__main__':
    init_db()
    
    # Get configuration from environment
    flask_env = os.getenv('FLASK_ENV', 'production')
    debug_mode = os.getenv('FLASK_DEBUG', 'False').lower() == 'true'
    host = os.getenv('HOST', '127.0.0.1')
    port = int(os.getenv('PORT', 5003))
    
    print("\n" + "="*70)
    print("🇪🇺  EU TRIP TRACKER - Version " + APP_VERSION)
    print("="*70)
    print(f"Environment: {flask_env}")
    print(f"Debug Mode: {debug_mode}")
    print(f"URL: http://{host}:{port}")
    print(f"Session Timeout: {CONFIG.get('SESSION_IDLE_TIMEOUT_MINUTES', 30)} minutes")
    print("="*70)
    
    if debug_mode:
        print("\nWARNING: Debug mode is enabled!")
        print("This should NEVER be used in production!\n")
    
    if not os.path.exists('.env'):
        print("\nWARNING: .env file not found!")
        print("Copy env_template.txt to .env and configure your settings.\n")
    
    print("="*70 + "\n")
    
    app.run(debug=debug_mode, host=host, port=port)
